from fastapi import APIRouter, Depends, HTTPException, status as estado, Request, Response
from fastapi.responses import FileResponse
from fastapi.security import OAuth2PasswordBearer
from sqlalchemy.orm import Session, lazyload
from jose import JWTError, jwt
from database import get_db
import models
import schemas
import auth
from typing import List, Optional
from datetime import datetime, timedelta
import os
import uuid
from services import facturacion
from services.whatsapp import send_whatsapp_message, format_ticket_message

router = APIRouter()

oauth2_scheme = OAuth2PasswordBearer(tokenUrl="api/auth/login")

def get_current_user(request: Request, token: str = Depends(oauth2_scheme), db: Session = Depends(get_db)):
    credentials_exception = HTTPException(
        status_code=estado.HTTP_401_UNAUTHORIZED,
        detail="Could not validate credentials",
        headers={"WWW-Authenticate": "Bearer"},
    )
    try:
        payload = jwt.decode(token, auth.SECRET_KEY, algorithms=[auth.ALGORITHM])
        nombre_usuario: str = payload.get("sub")
        if nombre_usuario is None:
            raise credentials_exception
        token_data = schemas.TokenData(nombre_usuario=nombre_usuario)
    except JWTError:
        raise credentials_exception
        
    user = db.query(models.Usuario).filter(models.Usuario.nombre_usuario == token_data.nombre_usuario).first()
    if user is None:
        raise credentials_exception
        
    # Verificar estado del Tenant si tiene uno
    if user.inquilino_id:
        inquilino = db.query(models.Inquilino).filter(models.Inquilino.id == user.inquilino_id).first()
        if not inquilino:
            raise HTTPException(
                status_code=estado.HTTP_403_FORBIDDEN,
                detail="Tienda no encontrada."
            )
        
        # Check if subscription has expired
        if inquilino.nivel_plan == 'premium' and inquilino.fin_suscripcion:
            try:
                now_str = datetime.utcnow().isoformat()
                if now_str > inquilino.fin_suscripcion and inquilino.estado_suscripcion == 'active':
                    inquilino.estado_suscripcion = 'suspended'
                    db.commit()
            except Exception as e:
                print("Error checking subscription expiration:", e)
                
        # If suspended, block all endpoints EXCEPT billing, superadmin, and basic auth information
        path = request.url.path
        if inquilino.estado_suscripcion not in ["active", "trialing"]:
            is_billing_route = "/billing" in path or "/superadmin" in path or "/auth/inquilino" in path or "/auth/users" in path
            if not is_billing_route:
                raise HTTPException(
                    status_code=estado.HTTP_403_FORBIDDEN,
                    detail="Suscripción inactiva. Por favor verifique sus pagos."
                )
            
    return user

@router.get("/auth/inquilino-branding/{subdominio}")
def get_tenant_branding(subdominio: str, db: Session = Depends(get_db)):
    inquilino = db.query(models.Inquilino).filter(models.Inquilino.subdominio == subdominio).first()
    if not inquilino:
        raise HTTPException(status_code=404, detail="Tienda no encontrada")
    
    plan = inquilino.nivel_plan or "free"
    estado_sub = inquilino.estado_suscripcion or "active"
    
    settings = db.query(models.ConfiguracionesTienda).filter(models.ConfiguracionesTienda.inquilino_id == inquilino.id).first()
    if not settings:
        return {
            "nombre_tienda": inquilino.nombre,
            "logo_url": None,
            "color_primario": "#064E3B",
            "color_secundario": "#DC2626",
            "nivel_plan": plan,
            "estado_suscripcion": estado_sub
        }
    
    return {
        "nombre_tienda": settings.nombre_tienda,
        "logo_url": settings.logo_url,
        "color_primario": settings.color_primario,
        "color_secundario": settings.color_secundario,
        "nivel_plan": plan,
        "estado_suscripcion": estado_sub
    }

@router.post("/auth/register-inquilino")
def register_tenant(req: schemas.InquilinoRegistroRequest, db: Session = Depends(get_db)):
    # 1. Validar si usuario admin ya existe
    existing_user = db.query(models.Usuario).filter(models.Usuario.nombre_usuario == req.usuario_admin).first()
    if existing_user:
        raise HTTPException(status_code=400, detail="El usuario administrador ya existe")

    # 2. Validar subdominio único si se provee
    if req.subdominio:
        existing_tenant = db.query(models.Inquilino).filter(models.Inquilino.subdominio == req.subdominio).first()
        if existing_tenant:
            raise HTTPException(status_code=400, detail="El subdominio ya está registrado")

    try:
        # 3. Crear el Tenant
        db_tenant = models.Inquilino(
            nombre=req.nombre_tienda,
            subdominio=req.subdominio or str(uuid.uuid4())[:8],
            estado_suscripcion="active",
            nivel_plan="free",  # Default to free plan
            creado_en=datetime.utcnow().isoformat()
        )
        db.add(db_tenant)
        db.commit()
        db.refresh(db_tenant)

        # 4. Crear el usuario Administrador para el Tenant
        contrasena_encriptada = auth.get_password_hash(req.contrasena_admin)
        db_user = models.Usuario(
            inquilino_id=db_tenant.id,
            nombre_usuario=req.usuario_admin,
            nombre_completo=req.nombre_admin,
            contrasena_encriptada=contrasena_encriptada,
            rol="admin"
        )
        db.add(db_user)

        # 5. Crear StoreSettings por defecto para este Tenant
        db_settings = models.ConfiguracionesTienda(
            inquilino_id=db_tenant.id,
            nombre_tienda=req.nombre_tienda,
            pie_ticket="¡Gracias por su compra!",
            logo_url=req.logo_url,
            color_primario=req.color_primario or "#064E3B",
            color_secundario=req.color_secundario or "#064E3B"
        )
        db.add(db_settings)

        db.commit()

        # Log action
        log_entry = models.BitacoraUsuario(
            inquilino_id=db_tenant.id,
            usuario=db_user.nombre_usuario,
            nombre_completo=db_user.nombre_completo,
            rol=db_user.rol,
            accion="creacion",
            detalles="Usuario administrador inicial creado durante el registro de la tienda.",
            fecha_hora=datetime.utcnow().isoformat()
        )
        db.add(log_entry)
        db.commit()

        return {
            "message": "Tenant and Admin created successfully", 
            "inquilino_id": db_tenant.id,
            "subdominio": db_tenant.subdominio
        }
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=400, detail=str(e))

@router.post("/auth/register")
def register(user: schemas.UsuarioCreate, db: Session = Depends(get_db), current_user: models.Usuario = Depends(get_current_user)):
    existing_user = db.query(models.Usuario).filter(models.Usuario.nombre_usuario == user.nombre_usuario).first()
    if existing_user:
        raise HTTPException(status_code=400, detail="El usuario ya existe")

    # Validar permisos
    if current_user.rol not in ['admin', 'supervisor']:
        raise HTTPException(status_code=403, detail="No tienes permisos suficientes para registrar usuarios")
        
    # Hash the password
    contrasena_encriptada = auth.get_password_hash(user.contrasena)
    
    db_user = models.Usuario(
        inquilino_id=current_user.inquilino_id,
        nombre_usuario=user.nombre_usuario, 
        nombre_completo=user.nombre_completo, 
        contrasena_encriptada=contrasena_encriptada,
        rol=user.rol
    )
    db.add(db_user)
    db.commit()
    db.refresh(db_user)

    # Log action
    log_entry = models.BitacoraUsuario(
        inquilino_id=current_user.inquilino_id,
        nombre_usuario=db_user.nombre_usuario,
        nombre_completo=db_user.nombre_completo,
        rol=db_user.rol,
        action="creacion",
        details=f"Usuario creado por {current_user.nombre_usuario} ({current_user.nombre_completo}) con el rol de {db_user.rol}.",
        fecha_hora=datetime.utcnow().isoformat()
    )
    db.add(log_entry)
    db.commit()
    
    return {"message": "User created successfully"}

@router.post("/auth/login", response_model=schemas.Token)
def login(user: schemas.UsuarioLogin, db: Session = Depends(get_db)):
    target_tenant = None
    if user.subdominio and user.subdominio.strip().lower() not in ["principal", "www", "localhost"]:
        target_tenant = db.query(models.Inquilino).filter(
            models.Inquilino.subdominio == user.subdominio.strip().lower()
        ).first()

    if target_tenant:
        # Buscar primero un usuario perteneciente exactamente a esta tienda
        db_user = db.query(models.Usuario).filter(
            models.Usuario.nombre_usuario == user.nombre_usuario.strip(),
            models.Usuario.inquilino_id == target_tenant.id
        ).first()

        # Si no existe en la tienda, permitir únicamente a administradores globales (Inquilino 1 con rol admin)
        if not db_user and target_tenant.id != 1:
            db_user = db.query(models.Usuario).filter(
                models.Usuario.nombre_usuario == user.nombre_usuario.strip(),
                models.Usuario.inquilino_id == 1,
                models.Usuario.rol == 'admin'
            ).first()

        if not db_user:
            raise HTTPException(
                status_code=estado.HTTP_401_UNAUTHORIZED,
                detail="Usuario o contraseña incorrectos",
                headers={"WWW-Authenticate": "Bearer"},
            )
    else:
        db_user = db.query(models.Usuario).filter(
            models.Usuario.nombre_usuario == user.nombre_usuario.strip()
        ).first()

    if not db_user or not auth.verify_password(user.contrasena, db_user.contrasena_encriptada):
        raise HTTPException(
            status_code=estado.HTTP_401_UNAUTHORIZED,
            detail="Usuario o contraseña incorrectos",
            headers={"WWW-Authenticate": "Bearer"},
        )

    # Verificar estado del Tenant si tiene uno
    sub_status = "active"
    subdominio = None
    if db_user.inquilino_id:
        inquilino = db.query(models.Inquilino).filter(models.Inquilino.id == db_user.inquilino_id).first()
        if not inquilino:
            raise HTTPException(status_code=estado.HTTP_403_FORBIDDEN, detail="Tienda no encontrada")
        
        # Check if subscription has expired
        if inquilino.nivel_plan == 'premium' and inquilino.fin_suscripcion:
            try:
                now_str = datetime.utcnow().isoformat()
                if now_str > inquilino.fin_suscripcion and inquilino.estado_suscripcion == 'active':
                    inquilino.estado_suscripcion = 'suspended'
                    db.commit()
            except Exception:
                pass
        sub_status = inquilino.estado_suscripcion
        subdominio = inquilino.subdominio
        
    access_token = auth.create_access_token(data={
        "sub": db_user.nombre_usuario, 
        "rol": db_user.rol,
        "inquilino_id": db_user.inquilino_id
    })
    
    return {
        "access_token": access_token, 
        "token_type": "bearer",
        "usuario": {
            "id": db_user.id, 
            "inquilino_id": db_user.inquilino_id,
            "nombre_usuario": db_user.nombre_usuario, 
            "nombre_completo": db_user.nombre_completo,
            "rol": db_user.rol,
            "estado_suscripcion": sub_status,
            "subdominio": subdominio
        }
    }

@router.get("/auth/inquilino", response_model=schemas.InquilinoResponse)
def get_current_tenant(db: Session = Depends(get_db), current_user: models.Usuario = Depends(get_current_user)):
    if not current_user.inquilino_id:
        raise HTTPException(status_code=404, detail="No se encontró tienda vinculada a tu usuario")
    inquilino = db.query(models.Inquilino).filter(models.Inquilino.id == current_user.inquilino_id).first()
    if not inquilino:
        raise HTTPException(status_code=404, detail="Tienda no encontrada")
    return inquilino

@router.post("/auth/inquilino/change-plan", response_model=schemas.InquilinoResponse)
def change_tenant_plan(req: schemas.InquilinoCambioPlanRequest, db: Session = Depends(get_db), current_user: models.Usuario = Depends(get_current_user)):
    if current_user.rol != 'admin':
        raise HTTPException(status_code=403, detail="Solo el administrador de la tienda puede cambiar de plan")
    if not current_user.inquilino_id:
        raise HTTPException(status_code=404, detail="No se encontró tienda vinculada a tu usuario")
    inquilino = db.query(models.Inquilino).filter(models.Inquilino.id == current_user.inquilino_id).first()
    if not inquilino:
        raise HTTPException(status_code=404, detail="Tienda no encontrada")
    
    inquilino.nivel_plan = req.nivel_plan
    db.commit()
    db.refresh(inquilino)
    return inquilino

@router.get("/auth/users", response_model=List[schemas.UsuarioResponse])
def get_users(db: Session = Depends(get_db), current_user: models.Usuario = Depends(get_current_user)):
    if current_user.rol != 'admin':
        raise HTTPException(status_code=403, detail="No tienes permisos suficientes")
    return db.query(models.Usuario).filter(models.Usuario.inquilino_id == current_user.inquilino_id).all()

@router.put("/auth/users/{usuario_id}", response_model=schemas.UsuarioResponse)
def update_user(usuario_id: int, user_data: schemas.UsuarioUpdate, db: Session = Depends(get_db), current_user: models.Usuario = Depends(get_current_user)):
    if current_user.rol != 'admin' and current_user.id != usuario_id:
        raise HTTPException(status_code=403, detail="No tienes permisos suficientes")
        
    db_user = db.query(models.Usuario).filter(
        models.Usuario.id == usuario_id,
        models.Usuario.inquilino_id == current_user.inquilino_id
    ).first()
    if not db_user:
        raise HTTPException(status_code=404, detail="Usuario no encontrado")
    
    update_data = user_data.dict(exclude_unset=True)
    if "contrasena" in update_data and update_data["contrasena"]:
        db_user.contrasena_encriptada = auth.get_password_hash(update_data["contrasena"])
        del update_data["contrasena"]
        
    for key, value in update_data.items():
        setattr(db_user, key, value)
        
    db.commit()
    db.refresh(db_user)

    # Log action
    log_entry = models.BitacoraUsuario(
        inquilino_id=current_user.inquilino_id,
        usuario=db_user.nombre_usuario,
        nombre_completo=db_user.nombre_completo,
        rol=db_user.rol,
        accion="actualizacion",
        detalles=f"Usuario modificado por {current_user.nombre_usuario} ({current_user.nombre_completo}).",
        fecha_hora=datetime.utcnow().isoformat()
    )
    db.add(log_entry)
    db.commit()
    
    return db_user

@router.delete("/auth/users/{usuario_id}")
def delete_user(usuario_id: int, db: Session = Depends(get_db), current_user: models.Usuario = Depends(get_current_user)):
    if current_user.rol != 'admin':
        raise HTTPException(status_code=403, detail="No tienes permisos suficientes")
        
    db_user = db.query(models.Usuario).filter(
        models.Usuario.id == usuario_id,
        models.Usuario.inquilino_id == current_user.inquilino_id
    ).first()
    if not db_user:
        raise HTTPException(status_code=404, detail="Usuario no encontrado")
        
    # Log action
    log_entry = models.BitacoraUsuario(
        inquilino_id=current_user.inquilino_id,
        usuario=db_user.nombre_usuario,
        nombre_completo=db_user.nombre_completo,
        rol=db_user.rol,
        accion="eliminacion",
        detalles=f"Usuario eliminado por {current_user.nombre_usuario} ({current_user.nombre_completo}).",
        fecha_hora=datetime.utcnow().isoformat()
    )
    db.add(log_entry)
    
    db.delete(db_user)
    db.commit()
    return {"message": "Usuario eliminado exitosamente"}

@router.get("/auth/users/logs", response_model=List[schemas.BitacoraUsuarioResponse])
def get_user_logs(db: Session = Depends(get_db), current_user: models.Usuario = Depends(get_current_user)):
    if current_user.rol not in ['admin', 'supervisor']:
        raise HTTPException(status_code=403, detail="Acceso denegado. Se requieren privilegios elevados.")
    
    logs = db.query(models.BitacoraUsuario).filter(
        models.BitacoraUsuario.inquilino_id == current_user.inquilino_id
    ).order_by(models.BitacoraUsuario.id.desc()).all()
    return logs

def search_product_image(query: str) -> Optional[str]:
    import urllib.request
    import urllib.parse
    import re
    import json
    
    query_clean = query.strip()
    if not query_clean:
        return None
        
    # 1. Try global OpenFoodFacts CGI search first (most accurate for grocery products)
    try:
        url = "https://world.openfoodfacts.org/cgi/search.pl?search_terms=" + urllib.parse.quote(query_clean) + "&search_simple=1&action=process&json=1"
        req = urllib.request.Request(
            url, 
            headers={'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36'}
        )
        with urllib.request.urlopen(req, timeout=2.0) as response:
            data = json.loads(response.read().decode('utf-8'))
        products = data.get("products", [])
        if products:
            for p in products:
                img_url = p.get("image_front_url") or p.get("image_url") or p.get("image_front_small_url")
                if img_url and img_url.startswith("http"):
                    return img_url
    except Exception:
        pass
        
    # 2. Fall back to Bing Images search with strict filters to retrieve real retail packaging
    try:
        search_query = query_clean
        # Auto-tune query for generic terms to get actual bottles/packaging
        if any(w in query_clean.lower() for w in ["salsa", "aceite", "leche", "crema", "agua", "refresco"]):
            search_query += " botella"
            
        url = "https://www.bing.com/images/search?q=" + urllib.parse.quote(search_query)
        req = urllib.request.Request(
            url, 
            headers={'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36'}
        )
        with urllib.request.urlopen(req, timeout=3.0) as response:
            html = response.read().decode('utf-8', errors='ignore')
            
        murls = re.findall(r'murl&quot;:&quot;(http[^&]+)&quot;', html)
        
        BLOCKED_DOMAINS = [
            "freepik.com", "shutterstock.com", "pinterest.com", "pinimg.com",
            "lovepik.com", "pngtree.com", "vecteezy.com", "depositphotos.com",
            "dreamstime.com", "123rf.com", "alamy.com", "canva.com",
            "vectorportal.com", "pixabay.com", "vectorstock.com", "istockphoto.com",
            "flaticon.com", "pngwing.com", "klipartz.com", "kindpng.com",
            "pngfind.com", "cleanpng.com", "pngegg.com", "favpng.com",
            "subpng.com", "pngsucai.com", "mksucai.com", "699pic.com",
            "freeimages.com", "pixy.org", "clipart", "gograph.com",
            "canstockphoto.com", "vector.me", "tnaflix.com", "xvideos.com",
            "pornhub.com", "redtube.com", "youporn.com", "xnxx.com", "spankbang.com",
            "eporner.com", "tube8.com", "porn.com"
        ]
        
        BLOCKED_KEYWORDS = [
            "logo", "icon", "vector", "dibujo", "clipart", "iluminacion", 
            "background", "fondo", "ilustration", "ilustracion", "banner", 
            "mockup", "silueta", "silhouette", "esquema", "porn", "tnaflix",
            "xxx", "adult", "sexy", "naked", "nude", "erotic"
        ]
        
        MOTOR_KEYWORDS = ["motor", "transmision", "transmission", "atf", "mobil", "lubricante", "refaccionaria", "mirefaccion"]
        
        for u in murls:
            u_lower = u.lower()
            blocked_domain = any(d in u_lower for d in BLOCKED_DOMAINS)
            blocked_kw = any(kw in u_lower for kw in BLOCKED_KEYWORDS)
            blocked_motor = any(mk in u_lower for mk in MOTOR_KEYWORDS)
            if not blocked_domain and not blocked_kw and not blocked_motor:
                return u
                
        img_srcs = re.findall(r'src="(http[^"]+)"', html)
        for src in img_srcs:
            src_lower = src.lower()
            if not any(x in src_lower for x in ["logo", "icon", "bing", "gstatic"]) and not any(d in src_lower for d in BLOCKED_DOMAINS) and not any(mk in src_lower for mk in MOTOR_KEYWORDS):
                return src
    except Exception:
        pass
        
    return None

@router.get("/inventory/search-imagen")
def get_product_image_search(q: str, db: Session = Depends(get_db), current_user: models.Usuario = Depends(get_current_user)):
    if not q or not q.strip():
        return {"image_url": None}
    img_url = search_product_image(q.strip())
    return {"image_url": img_url}

@router.get("/inventory/", response_model=List[schemas.ProductoResponse])
def get_inventory(db: Session = Depends(get_db), current_user: models.Usuario = Depends(get_current_user)):
    products = db.query(models.Producto).filter(models.Producto.inquilino_id == current_user.inquilino_id).all()
    return products

@router.post("/inventory/", response_model=schemas.ProductoResponse)
def create_product(producto: schemas.ProductoCreate, db: Session = Depends(get_db), current_user: models.Usuario = Depends(get_current_user)):
    if current_user.rol not in ['admin', 'supervisor']:
        raise HTTPException(status_code=403, detail="No tienes permisos suficientes")
    
    # Validar límite de plan gratis
    inquilino = db.query(models.Inquilino).filter(models.Inquilino.id == current_user.inquilino_id).first()
    if inquilino and inquilino.nivel_plan == 'free':
        prod_count = db.query(models.Producto).filter(models.Producto.inquilino_id == current_user.inquilino_id).count()
        if prod_count >= 50:
            raise HTTPException(
                status_code=400,
                detail="Límite del plan gratuito alcanzado (máximo 50 productos). Por favor, actualiza tu cuenta a Premium."
            )

    prod_data = producto.dict()
    variants_data = prod_data.pop("variantes", [])
    
    db_product = models.Producto(**prod_data, inquilino_id=current_user.inquilino_id)
    db.add(db_product)
    db.commit()
    db.refresh(db_product)
    
    if variants_data:
        for v in variants_data:
            db_var = models.VarianteProducto(
                inquilino_id=current_user.inquilino_id,
                producto_id=db_product.id,
                nombre=v["nombre"],
                codigo_barras=v.get("codigo_barras"),
                precio_costo=v.get("precio_costo") if v.get("precio_costo") is not None else db_product.precio_costo,
                precio=v.get("precio") if v.get("precio") is not None else db_product.precio,
                cantidad=v["cantidad"]
            )
            db.add(db_var)
        db.commit()
        db.refresh(db_product)
    return db_product

@router.put("/inventory/{producto_id}", response_model=schemas.ProductoResponse)
def update_product(producto_id: int, producto: schemas.ProductoCreate, db: Session = Depends(get_db), current_user: models.Usuario = Depends(get_current_user)):
    if current_user.rol not in ['admin', 'supervisor']:
        raise HTTPException(status_code=403, detail="No tienes permisos suficientes")
    db_product = db.query(models.Producto).filter(
        models.Producto.id == producto_id,
        models.Producto.inquilino_id == current_user.inquilino_id
    ).first()
    if not db_product:
        raise HTTPException(status_code=404, detail="Product not found")
    
    prod_data = producto.dict()
    variants_data = prod_data.pop("variantes", [])
    
    for key, value in prod_data.items():
        setattr(db_product, key, value)
    
    # Sync variantes (simple delete and recreate)
    db.query(models.VarianteProducto).filter(
        models.VarianteProducto.producto_id == producto_id,
        models.VarianteProducto.inquilino_id == current_user.inquilino_id
    ).delete()
    if variants_data:
        for v in variants_data:
            db_var = models.VarianteProducto(
                inquilino_id=current_user.inquilino_id,
                producto_id=db_product.id,
                nombre=v["nombre"],
                codigo_barras=v.get("codigo_barras"),
                precio_costo=v.get("precio_costo") if v.get("precio_costo") is not None else db_product.precio_costo,
                precio=v.get("precio") if v.get("precio") is not None else db_product.precio,
                cantidad=v["cantidad"]
            )
            db.add(db_var)
            
    db.commit()
    db.refresh(db_product)
    return db_product

@router.delete("/inventory/{producto_id}")
def delete_product(producto_id: int, db: Session = Depends(get_db), current_user: models.Usuario = Depends(get_current_user)):
    if current_user.rol not in ['admin', 'supervisor']:
        raise HTTPException(status_code=403, detail="No tienes permisos suficientes")
    db_product = db.query(models.Producto).filter(
        models.Producto.id == producto_id,
        models.Producto.inquilino_id == current_user.inquilino_id
    ).first()
    if not db_product:
        raise HTTPException(status_code=404, detail="Product not found")
    
    db.delete(db_product)
    db.commit()
    return {"message": "Product deleted successfully"}

@router.post("/inventory/{producto_id}/sell", response_model=schemas.ProductoResponse)
def sell_product(producto_id: int, sell_data: schemas.ProductoVenta, db: Session = Depends(get_db), current_user: models.Usuario = Depends(get_current_user)):
    db_product = db.query(models.Producto).filter(
        models.Producto.id == producto_id,
        models.Producto.inquilino_id == current_user.inquilino_id
    ).first()
    if not db_product:
        raise HTTPException(status_code=404, detail="Product not found")
        
    from sqlalchemy import text
    try:
        db.execute(
            text("SELECT vender_producto(:producto_id, :cantidad, :creado_en)"),
            {
                "producto_id": producto_id,
                "cantidad": sell_data.cantidad,
                "creado_en": datetime.utcnow().isoformat()
            }
        )
        db.commit()
        
        # Actualizar inquilino_id en el registro de venta recién creado por el SP vender_producto
        latest_sale = db.query(models.HistorialVenta).filter(
            models.HistorialVenta.producto_id == producto_id,
            models.HistorialVenta.inquilino_id == None
        ).order_by(models.HistorialVenta.id.desc()).first()
        if latest_sale:
            latest_sale.inquilino_id = current_user.inquilino_id
            latest_sale.usuario_id = current_user.id
            db.commit()
            
    except Exception as e:
        db.rollback()
        # Parse error message to show a clean response
        error_msg = str(e)
        if hasattr(e, 'orig') and e.orig:
            orig_msg = str(e.orig)
            if "ERROR:" in orig_msg:
                parts = orig_msg.split("ERROR:")
                if len(parts) > 1:
                    error_msg = parts[1].split("\n")[0].strip()
            else:
                error_msg = orig_msg.split("\n")[0].strip()
        raise HTTPException(status_code=400, detail=error_msg)
        
    db_product = db.query(models.Producto).filter(
        models.Producto.id == producto_id,
        models.Producto.inquilino_id == current_user.inquilino_id
    ).first()
    return db_product

@router.get("/inventory/sales_report")
def get_sales_report(db: Session = Depends(get_db), current_user: models.Usuario = Depends(get_current_user)):
    if current_user.rol not in ['admin', 'supervisor']:
        raise HTTPException(status_code=403, detail="No tienes permisos suficientes")
    products = db.query(models.Producto).filter(models.Producto.inquilino_id == current_user.inquilino_id).all()
    sales = db.query(models.HistorialVenta).filter(
        models.HistorialVenta.inquilino_id == current_user.inquilino_id,
        models.HistorialVenta.cancelado == False
    ).all()
    
    now = datetime.utcnow()
    today_start = now.replace(hour=0, minute=0, second=0, microsecond=0)
    week_start = today_start - timedelta(days=now.weekday()) # Monday of current week
    month_start = today_start.replace(day=1)
    
    report = []
    for p in products:
        p_sales = [s for s in sales if s.producto_id == p.id]
        
        sales_today = sum(s.cantidad for s in p_sales if datetime.fromisoformat(s.creado_en) >= today_start)
        sales_week = sum(s.cantidad for s in p_sales if datetime.fromisoformat(s.creado_en) >= week_start)
        sales_month = sum(s.cantidad for s in p_sales if datetime.fromisoformat(s.creado_en) >= month_start)
        
        def get_sale_revenue(s):
            precio = s.price_sold if s.price_sold is not None else p.precio
            return (precio * s.cantidad) - s.discount
 
        revenue_today = sum(get_sale_revenue(s) for s in p_sales if datetime.fromisoformat(s.creado_en) >= today_start)
        revenue_week = sum(get_sale_revenue(s) for s in p_sales if datetime.fromisoformat(s.creado_en) >= week_start)
        revenue_month = sum(get_sale_revenue(s) for s in p_sales if datetime.fromisoformat(s.creado_en) >= month_start)
        revenue_total = sum(get_sale_revenue(s) for s in p_sales)
        
        report.append({
            "id": p.id,
            "nombre": p.nombre,
            "precio": p.precio,
            "cantidad": p.cantidad,
            "sold_total": p.vendido,
            "sales_today": sales_today,
            "revenue_today": revenue_today,
            "sales_week": sales_week,
            "revenue_week": revenue_week,
            "sales_month": sales_month,
            "revenue_month": revenue_month,
            "revenue_total": revenue_total
        })
        
    return report

@router.get("/dashboard/stats")
def get_stats(db: Session = Depends(get_db), current_user: models.Usuario = Depends(get_current_user)):
    if current_user.rol != 'admin':
        raise HTTPException(status_code=403, detail="No tienes permisos suficientes")
    products = db.query(models.Producto).filter(models.Producto.inquilino_id == current_user.inquilino_id).all()
    total_stock = sum(p.cantidad for p in products)
    total_sold = sum(p.vendido for p in products)
    low_stock = sum(1 for p in products if p.cantidad < 20)
    
    return {
        "total_stock": total_stock,
        "total_sold": total_sold,
        "low_stock_alerts": low_stock
    }

@router.get("/sales/recent")
def get_recent_sales(db: Session = Depends(get_db), current_user: models.Usuario = Depends(get_current_user)):
    results = db.query(
        models.HistorialVenta.id,
        models.HistorialVenta.producto_id,
        models.HistorialVenta.cantidad,
        models.HistorialVenta.creado_en,
        models.HistorialVenta.cancelado,
        models.HistorialVenta.motivo_cancelacion,
        models.HistorialVenta.autorizado_por,
        models.Producto.nombre.label("nombre_producto"),
        models.Producto.precio.label("product_price"),
        models.Usuario.nombre_completo.label("cashier_name")
    ).join(
        models.Producto, models.Producto.id == models.HistorialVenta.producto_id
    ).outerjoin(
        models.Usuario, models.Usuario.id == models.HistorialVenta.usuario_id
    ).filter(
        models.HistorialVenta.inquilino_id == current_user.inquilino_id
    ).order_by(
        models.HistorialVenta.id.desc()
    ).limit(50).all()
    
    return [
        {
            "id": r.id,
            "producto_id": r.producto_id,
            "nombre_producto": r.nombre_producto,
            "product_price": r.product_price,
            "cantidad": r.cantidad,
            "creado_en": r.creado_en,
            "cancelado": r.cancelado,
            "motivo_cancelacion": r.motivo_cancelacion,
            "autorizado_por": r.autorizado_por,
            "cashier_name": r.cashier_name or "Desconocido"
        }
        for r in results
    ]

@router.post("/sales/{venta_id}/cancel")
def cancel_sale_endpoint(venta_id: int, cancel_data: schemas.CancelarVentaRequest, db: Session = Depends(get_db), current_user: models.Usuario = Depends(get_current_user)):
    authorized_by_username = current_user.nombre_usuario
    auth_user = cancel_data.auth_username or cancel_data.usuario_autorizacion
    auth_pass = cancel_data.auth_password or cancel_data.contrasena_autorizacion

    if current_user.rol == "cajero" or current_user.rol == "user":
        # Requiere credenciales de admin o supervisor
        if not auth_user or not auth_pass:
            raise HTTPException(status_code=403, detail="Las cancelaciones están restringidas para cajeros. Se requieren credenciales de Supervisor o Administrador.")
        
        supervisor_user = db.query(models.Usuario).filter(
            models.Usuario.nombre_usuario == auth_user,
            models.Usuario.inquilino_id == current_user.inquilino_id
        ).first()
        if not supervisor_user or not auth.verify_password(auth_pass, supervisor_user.contrasena_encriptada):
            raise HTTPException(status_code=403, detail="Usuario o contraseña del supervisor incorrectos.")
        
        if supervisor_user.rol not in ["admin", "supervisor"]:
            raise HTTPException(status_code=403, detail="El usuario autorizador no tiene permisos de Supervisor o Administrador.")
        authorized_by_username = supervisor_user.nombre_usuario

    # Cancelar venta transaccional en python
    try:
        sale = db.query(models.HistorialVenta).filter(
            models.HistorialVenta.id == venta_id,
            models.HistorialVenta.inquilino_id == current_user.inquilino_id
        ).with_for_update().first()
        
        if not sale:
            raise HTTPException(status_code=404, detail="La venta no existe.")
            
        if sale.cancelado:
            raise HTTPException(status_code=400, detail="Esta venta ya fue cancelada o devuelta por completo.")
            
        # Determinar cantidad a cancelar/devolver
        qty_to_cancel = cancel_data.cantidad
        if qty_to_cancel is None:
            qty_to_cancel = sale.cantidad
            
        if qty_to_cancel <= 0 or qty_to_cancel > sale.cantidad:
            raise HTTPException(status_code=400, detail=f"Cantidad inválida a cancelar. Disponible: {sale.cantidad}, Solicitado: {qty_to_cancel}")
            
        # Devolver stock
        if sale.variante_id:
            var = db.query(models.VarianteProducto).filter(
                models.VarianteProducto.id == sale.variante_id,
                models.VarianteProducto.inquilino_id == current_user.inquilino_id
            ).with_for_update().first()
            if var:
                var.cantidad += qty_to_cancel
                var.vendido -= qty_to_cancel
            prod = db.query(models.Producto).filter(
                models.Producto.id == sale.producto_id,
                models.Producto.inquilino_id == current_user.inquilino_id
            ).options(lazyload(models.Producto.variantes)).with_for_update().first()
            if prod:
                prod.cantidad += qty_to_cancel
                prod.vendido -= qty_to_cancel
        else:
            prod = db.query(models.Producto).filter(
                models.Producto.id == sale.producto_id,
                models.Producto.inquilino_id == current_user.inquilino_id
            ).options(lazyload(models.Producto.variantes)).with_for_update().first()
            if prod:
                prod.cantidad += qty_to_cancel
                prod.vendido -= qty_to_cancel
                
        # Calcular reembolsos proporcionales y deducir del turno de caja activo
        ratio = qty_to_cancel / sale.cantidad
        discount_refund = sale.discount * ratio
        cash_refund = (sale.cash_amount or 0.0) * ratio
        card_refund = (sale.card_amount or 0.0) * ratio
        
        if sale.turno_id:
            shift = db.query(models.Turno).filter(
                models.Turno.id == sale.turno_id,
                models.Turno.inquilino_id == current_user.inquilino_id,
                models.Turno.estado == "open"
            ).first()
            if shift:
                cash_to_deduct = 0.0
                if sale.payment_method == "efectivo":
                    cash_to_deduct = (sale.price_sold * qty_to_cancel) - discount_refund
                elif sale.payment_method == "mixto":
                    cash_to_deduct = cash_refund
                    
                shift.efectivo_final_esperado -= cash_to_deduct
                
        price_val = sale.price_sold if sale.price_sold is not None else 0.0
        
        # Guardar registro en product_returns
        prod_return = models.DevolucionProducto(
            inquilino_id=current_user.inquilino_id,
            venta_id=sale.id,
            producto_id=sale.producto_id,
            cantidad=qty_to_cancel,
            precio=price_val,
            motivo=cancel_data.motivo,
            autorizado_por=authorized_by_username,
            creado_en=datetime.utcnow().isoformat()
        )
        db.add(prod_return)
        
        # Actualizar el registro original de la venta
        sale.autorizado_por = authorized_by_username
        if qty_to_cancel == sale.cantidad:
            sale.cancelado = True
            sale.motivo_cancelacion = cancel_data.motivo
        else:
            sale.cantidad -= qty_to_cancel
            sale.discount -= discount_refund
            if sale.cash_amount:
                sale.cash_amount -= cash_refund
            if sale.card_amount:
                sale.card_amount -= card_refund
            
            partial_reason = f"Devolución parcial de {qty_to_cancel} pzs: {cancel_data.motivo}"
            if sale.motivo_cancelacion:
                sale.motivo_cancelacion += f" | {partial_reason}"
            else:
                sale.motivo_cancelacion = partial_reason
        
        db.commit()
        return {"message": f"Devolución de {qty_to_cancel} piezas procesada exitosamente."}
        
    except HTTPException:
        db.rollback()
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=400, detail=str(e))

@router.get("/inventory/returns_report", response_model=List[schemas.DevolucionResponse])
def get_returns_report(db: Session = Depends(get_db), current_user: models.Usuario = Depends(get_current_user)):
    if current_user.rol not in ['admin', 'supervisor']:
        raise HTTPException(status_code=403, detail="No tienes permisos suficientes")
    
    results = db.query(
        models.DevolucionProducto.id,
        models.DevolucionProducto.venta_id,
        models.DevolucionProducto.producto_id,
        models.DevolucionProducto.cantidad,
        models.DevolucionProducto.precio,
        models.DevolucionProducto.motivo,
        models.DevolucionProducto.autorizado_por,
        models.DevolucionProducto.creado_en,
        models.Producto.nombre.label("nombre_producto")
    ).join(
        models.Producto, models.Producto.id == models.DevolucionProducto.producto_id
    ).filter(
        models.DevolucionProducto.inquilino_id == current_user.inquilino_id
    ).order_by(
        models.DevolucionProducto.id.desc()
    ).all()
    
    return [
        {
            "id": r.id,
            "venta_id": r.venta_id,
            "producto_id": r.producto_id,
            "nombre_producto": r.nombre_producto,
            "cantidad": r.cantidad,
            "precio": r.precio,
            "motivo": r.motivo,
            "autorizado_por": r.autorizado_por,
            "creado_en": r.creado_en
        }
        for r in results
    ]

# --- NUEVOS ENDPOINTS MEJORAS POS ---

@router.post("/sales/checkout")
def checkout(checkout_data: schemas.PeticionCheckout, db: Session = Depends(get_db), current_user: models.Usuario = Depends(get_current_user)):
    shift = None
    if checkout_data.turno_id is not None:
        shift = db.query(models.Turno).filter(
            models.Turno.id == checkout_data.turno_id,
            models.Turno.inquilino_id == current_user.inquilino_id,
            models.Turno.estado == "open"
        ).first()
        if not shift and current_user.rol != 'admin':
            raise HTTPException(status_code=400, detail="No hay un turno de caja activo o el ID del turno es incorrecto.")
    elif current_user.rol != 'admin':
        raise HTTPException(status_code=400, detail="Se requiere un turno de caja activo para procesar la venta.")
        
    # Validar cliente si es pago a crédito
    customer = None
    if checkout_data.metodo_pago == "credito":
        if not checkout_data.cliente_id:
            raise HTTPException(status_code=400, detail="Debe seleccionar un cliente para realizar una venta a crédito.")
        customer = db.query(models.Cliente).filter(
            models.Cliente.id == checkout_data.cliente_id,
            models.Cliente.inquilino_id == current_user.inquilino_id
        ).with_for_update().first()
        if not customer:
            raise HTTPException(status_code=404, detail="El cliente seleccionado no existe.")
    elif checkout_data.cliente_id:
        customer = db.query(models.Cliente).filter(
            models.Cliente.id == checkout_data.cliente_id,
            models.Cliente.inquilino_id == current_user.inquilino_id
        ).first()

    try:
        now_str = datetime.utcnow().isoformat()
        subtotal = 0.0
        items_to_process = []
        
        for item in checkout_data.elementos:
            if item.variante_id:
                var = db.query(models.VarianteProducto).filter(
                    models.VarianteProducto.id == item.variante_id,
                    models.VarianteProducto.inquilino_id == current_user.inquilino_id
                ).with_for_update().first()
                if not var:
                    raise HTTPException(status_code=404, detail=f"La variante con ID {item.variante_id} no existe.")
                if var.cantidad < item.cantidad:
                    raise HTTPException(status_code=400, detail=f"Stock insuficiente para la variante {var.nombre}. Disponible: {var.cantidad}, Solicitado: {item.cantidad}")
                
                prod = db.query(models.Producto).filter(
                    models.Producto.id == var.producto_id,
                    models.Producto.inquilino_id == current_user.inquilino_id
                ).first()
                precio = var.precio if var.precio is not None else prod.precio
                cost = var.precio_costo if var.precio_costo is not None else prod.precio_costo
                subtotal += precio * item.cantidad
                items_to_process.append((item, var, prod, precio, cost))
            else:
                prod = db.query(models.Producto).filter(
                    models.Producto.id == item.producto_id,
                    models.Producto.inquilino_id == current_user.inquilino_id
                ).options(lazyload(models.Producto.variantes)).with_for_update().first()
                if not prod:
                    raise HTTPException(status_code=404, detail=f"El producto con ID {item.producto_id} no existe.")
                if prod.cantidad < item.cantidad:
                    raise HTTPException(status_code=400, detail=f"Stock insuficiente para el producto {prod.nombre}. Disponible: {prod.cantidad}, Solicitado: {item.cantidad}")
                
                precio = prod.precio
                cost = prod.precio_costo
                subtotal += precio * item.cantidad
                items_to_process.append((item, None, prod, precio, cost))
                
        if subtotal <= 0:
            raise HTTPException(status_code=400, detail="El total de la venta debe ser mayor a 0.")

        sale_total = subtotal - checkout_data.descuento

        # Validar límite de crédito del cliente si aplica
        if checkout_data.metodo_pago == "credito" and customer:
            if customer.saldo_actual + sale_total > customer.limite_credito:
                raise HTTPException(
                    status_code=400, 
                    detail=f"Límite de crédito excedido. Disponible: ${customer.limite_credito - customer.saldo_actual:.2f}, Total venta: ${sale_total:.2f}"
                )
            customer.saldo_actual += sale_total

        discount_ratio = checkout_data.descuento / subtotal if checkout_data.descuento > 0 else 0.0
        
        for item_data in items_to_process:
            item, var, prod, precio, cost = item_data
            qty = item.cantidad
            
            item_subtotal = precio * qty
            item_discount = item_subtotal * discount_ratio
            
            if var:
                var.cantidad -= qty
                var.vendido += qty
                prod.cantidad -= qty
                prod.vendido += qty
            else:
                prod.cantidad -= qty
                prod.vendido += qty
                
            sale_record = models.HistorialVenta(
                inquilino_id=current_user.inquilino_id,
                producto_id=prod.id,
                variante_id=var.id if var else None,
                turno_id=shift.id if shift else None,
                usuario_id=current_user.id,
                cantidad=qty,
                price_sold=precio,
                cost_price_sold=cost,
                discount=item_discount,
                payment_method=checkout_data.metodo_pago,
                cash_amount=checkout_data.monto_efectivo * (item_subtotal / subtotal) if checkout_data.metodo_pago == "mixto" else (sale_total if checkout_data.metodo_pago == "efectivo" else 0.0),
                card_amount=checkout_data.monto_tarjeta * (item_subtotal / subtotal) if checkout_data.metodo_pago == "mixto" else (sale_total if checkout_data.metodo_pago == "tarjeta" else 0.0),
                creado_en=now_str,
                cliente_id=checkout_data.cliente_id
            )
            db.add(sale_record)
            
        if shift:
            cash_sale_total = 0.0
            if checkout_data.metodo_pago == "efectivo":
                cash_sale_total = sale_total
            elif checkout_data.metodo_pago == "mixto":
                cash_sale_total = checkout_data.monto_efectivo
                
            shift.efectivo_final_esperado += cash_sale_total
        
        db.commit()
        db.refresh(sale_record)
        return {"message": "Venta procesada exitosamente", "venta_id": sale_record.id}

        
    except Exception as e:
        db.rollback()
        if isinstance(e, HTTPException):
            raise e
        raise HTTPException(status_code=400, detail=str(e))

@router.get("/shifts/active", response_model=Optional[schemas.TurnoResponse])
def get_active_shift(db: Session = Depends(get_db), current_user: models.Usuario = Depends(get_current_user)):
    shift = db.query(models.Turno).filter(
        models.Turno.usuario_id == current_user.id,
        models.Turno.inquilino_id == current_user.inquilino_id,
        models.Turno.estado == "open"
    ).first()
    return shift

@router.post("/shifts/open", response_model=schemas.TurnoResponse)
def open_shift(shift_data: schemas.TurnoCreate, db: Session = Depends(get_db), current_user: models.Usuario = Depends(get_current_user)):
    existing = db.query(models.Turno).filter(
        models.Turno.usuario_id == current_user.id,
        models.Turno.inquilino_id == current_user.inquilino_id,
        models.Turno.estado == "open"
    ).first()
    if existing:
        raise HTTPException(status_code=400, detail="Ya tienes un turno de caja abierto.")
        
    new_shift = models.Turno(
        inquilino_id=current_user.inquilino_id,
        usuario_id=current_user.id,
        hora_inicio=datetime.utcnow().isoformat(),
        efectivo_inicial=shift_data.efectivo_inicial,
        efectivo_final_esperado=shift_data.efectivo_inicial,
        estado="open"
    )
    db.add(new_shift)
    db.commit()
    db.refresh(new_shift)
    return new_shift

@router.post("/shifts/close", response_model=schemas.TurnoResponse)
def close_shift(close_data: schemas.TurnoClose, db: Session = Depends(get_db), current_user: models.Usuario = Depends(get_current_user)):
    shift = db.query(models.Turno).filter(
        models.Turno.usuario_id == current_user.id,
        models.Turno.inquilino_id == current_user.inquilino_id,
        models.Turno.estado == "open"
    ).first()
    if not shift:
        raise HTTPException(status_code=404, detail="No tienes ningún turno de caja activo.")
        
    shift.hora_fin = datetime.utcnow().isoformat()
    shift.efectivo_final_real = close_data.efectivo_final_real
    shift.diferencia = close_data.efectivo_final_real - shift.efectivo_final_esperado
    shift.estado = "closed"
    
    db.commit()
    db.refresh(shift)
    return shift

@router.post("/shifts/movement", response_model=schemas.MovimientoCajaResponse)
def add_cash_movement(movement: schemas.MovimientoCajaCreate, db: Session = Depends(get_db), current_user: models.Usuario = Depends(get_current_user)):
    shift = db.query(models.Turno).filter(
        models.Turno.usuario_id == current_user.id,
        models.Turno.inquilino_id == current_user.inquilino_id,
        models.Turno.estado == "open"
    ).first()
    if not shift:
        raise HTTPException(status_code=400, detail="No tienes un turno de caja abierto para registrar movimientos.")
        
    db_mov = models.MovimientoCaja(
        inquilino_id=current_user.inquilino_id,
        turno_id=shift.id,
        tipo=movement.tipo,
        monto=movement.monto,
        motivo=movement.motivo,
        creado_en=datetime.utcnow().isoformat()
    )
    
    if movement.tipo == "entrada":
        shift.efectivo_final_esperado += movement.monto
    elif movement.tipo in ["salida", "retiro_parcial"]:
        if shift.efectivo_final_esperado < movement.monto:
            raise HTTPException(status_code=400, detail="No puedes retirar una cantidad mayor a la que hay en caja actualmente.")
        shift.efectivo_final_esperado -= movement.monto
        
    db.add(db_mov)
    db.commit()
    db.refresh(db_mov)
    return db_mov

@router.get("/shifts/active-all")
def get_all_active_shifts(db: Session = Depends(get_db), current_user: models.Usuario = Depends(get_current_user)):
    if current_user.rol != 'admin':
        raise HTTPException(status_code=403, detail="No tienes permisos suficientes.")
        
    active_shifts = db.query(models.Turno).filter(
        models.Turno.estado == "open",
        models.Turno.inquilino_id == current_user.inquilino_id
    ).all()
    
    result = []
    for s in active_shifts:
        user = db.query(models.Usuario).filter(
            models.Usuario.id == s.usuario_id,
            models.Usuario.inquilino_id == current_user.inquilino_id
        ).first()
        result.append({
            "id": s.id,
            "usuario_id": s.usuario_id,
            "nombre_usuario": user.nombre_usuario if user else "Desconocido",
            "nombre_completo": user.nombre_completo if user else "Desconocido",
            "hora_inicio": s.hora_inicio,
            "efectivo_inicial": s.efectivo_inicial,
            "efectivo_final_esperado": s.efectivo_final_esperado,
            "estado": s.estado
        })
    return result

@router.post("/shifts/{turno_id}/close", response_model=schemas.TurnoResponse)
def close_any_shift(turno_id: int, close_data: schemas.TurnoClose, db: Session = Depends(get_db), current_user: models.Usuario = Depends(get_current_user)):
    if current_user.rol != 'admin':
        raise HTTPException(status_code=403, detail="No tienes permisos suficientes.")
        
    shift = db.query(models.Turno).filter(
        models.Turno.id == turno_id,
        models.Turno.inquilino_id == current_user.inquilino_id,
        models.Turno.estado == "open"
    ).first()
    if not shift:
        raise HTTPException(status_code=404, detail="Turno no encontrado o ya está cerrado.")
        
    shift.hora_fin = datetime.utcnow().isoformat()
    shift.efectivo_final_real = close_data.efectivo_final_real
    shift.diferencia = close_data.efectivo_final_real - shift.efectivo_final_esperado
    shift.estado = "closed"
    db.commit()
    db.refresh(shift)
    return shift

@router.get("/shifts/{turno_id}/report")
def get_shift_report(turno_id: int, db: Session = Depends(get_db), current_user: models.Usuario = Depends(get_current_user)):
    shift = db.query(models.Turno).filter(
        models.Turno.id == turno_id,
        models.Turno.inquilino_id == current_user.inquilino_id
    ).first()
    if not shift:
        raise HTTPException(status_code=404, detail="Turno no encontrado.")
        
    if current_user.rol == "cajero" and shift.usuario_id != current_user.id:
        raise HTTPException(status_code=403, detail="No tienes permisos para ver el reporte de otros turnos.")
        
    movements = db.query(models.MovimientoCaja).filter(
        models.MovimientoCaja.turno_id == shift.id,
        models.MovimientoCaja.inquilino_id == current_user.inquilino_id
    ).all()
    
    sales = db.query(
        models.HistorialVenta.id,
        models.HistorialVenta.cantidad,
        models.HistorialVenta.price_sold,
        models.HistorialVenta.discount,
        models.HistorialVenta.payment_method,
        models.HistorialVenta.cash_amount,
        models.HistorialVenta.card_amount,
        models.HistorialVenta.cancelado,
        models.Producto.nombre.label("nombre_producto")
    ).join(
        models.Producto, models.Producto.id == models.HistorialVenta.producto_id
    ).filter(
        models.HistorialVenta.turno_id == shift.id,
        models.HistorialVenta.inquilino_id == current_user.inquilino_id
    ).all()
    
    cash_sales = 0.0
    card_sales = 0.0
    credit_sales = 0.0
    cancelled_sales_total = 0.0
    
    sales_list = []
    for s in sales:
        total = (s.price_sold * s.cantidad) - s.discount
        if s.cancelado:
            cancelled_sales_total += total
        else:
            if s.payment_method == "efectivo":
                cash_sales += total
            elif s.payment_method == "tarjeta":
                card_sales += total
            elif s.payment_method == "credito":
                credit_sales += total
            elif s.payment_method == "mixto":
                cash_sales += s.cash_amount
                card_sales += s.card_amount
                
        sales_list.append({
            "id": s.id,
            "nombre_producto": s.nombre_producto,
            "cantidad": s.cantidad,
            "total": total,
            "payment_method": s.payment_method,
            "cancelado": s.cancelado
        })
        
    entries = sum(m.monto for m in movements if m.tipo == "entrada")
    withdrawals = sum(m.monto for m in movements if m.tipo in ["salida", "retiro_parcial"])
    
    cajero = db.query(models.Usuario).filter(
        models.Usuario.id == shift.usuario_id,
        models.Usuario.inquilino_id == current_user.inquilino_id
    ).first()
    
    return {
        "shift": {
            "id": shift.id,
            "cashier_name": cajero.nombre_completo if cajero else "Desconocido",
            "hora_inicio": shift.hora_inicio,
            "hora_fin": shift.hora_fin,
            "efectivo_inicial": shift.efectivo_inicial,
            "efectivo_final_real": shift.efectivo_final_real,
            "efectivo_final_esperado": shift.efectivo_final_esperado,
            "diferencia": shift.diferencia,
            "estado": shift.estado
        },
        "totals": {
            "cash_sales": cash_sales,
            "card_sales": card_sales,
            "credit_sales": credit_sales,
            "total_sales": cash_sales + card_sales + credit_sales,
            "cash_entries": entries,
            "cash_withdrawals": withdrawals,
            "cancelled_sales_total": cancelled_sales_total
        },
        "movements": [
            {
                "id": m.id,
                "tipo": m.tipo,
                "monto": m.monto,
                "motivo": m.motivo,
                "creado_en": m.creado_en
            }
            for m in movements
        ],
        "sales": sales_list
    }

@router.get("/reports/profit-margin")
def get_profit_margin_report(db: Session = Depends(get_db), current_user: models.Usuario = Depends(get_current_user)):
    if current_user.rol not in ["admin", "supervisor"]:
        raise HTTPException(status_code=403, detail="No tienes permisos suficientes.")
        
    products = db.query(models.Producto).filter(models.Producto.inquilino_id == current_user.inquilino_id).all()
    sales = db.query(models.HistorialVenta).filter(
        models.HistorialVenta.inquilino_id == current_user.inquilino_id,
        models.HistorialVenta.cancelado == False
    ).all()
    
    report = []
    total_revenue = 0.0
    costo_total = 0.0
    
    for p in products:
        p_sales = [s for s in sales if s.producto_id == p.id]
        qty_sold = sum(s.cantidad for s in p_sales)
        
        def get_sale_revenue(s):
            precio = s.price_sold if s.price_sold is not None else p.precio
            return (precio * s.cantidad) - s.discount
            
        def get_sale_cost(s):
            cost = s.cost_price_sold if s.cost_price_sold is not None else p.precio_costo
            return cost * s.cantidad

        revenue = sum(get_sale_revenue(s) for s in p_sales)
        cost = sum(get_sale_cost(s) for s in p_sales)
        
        profit = revenue - cost
        margin_pct = (profit / revenue * 100) if revenue > 0 else 0.0
        
        total_revenue += revenue
        costo_total += cost
        
        report.append({
            "producto_id": p.id,
            "nombre_producto": p.nombre,
            "quantity_sold": qty_sold,
            "revenue": revenue,
            "cost": cost,
            "profit": profit,
            "margin_percentage": margin_pct
        })
        
    report = sorted(report, key=lambda x: x["quantity_sold"], reverse=True)
    
    return {
        "summary": {
            "total_revenue": total_revenue,
            "costo_total": costo_total,
            "total_profit": total_revenue - costo_total,
            "average_margin_percentage": ((total_revenue - costo_total) / total_revenue * 100) if total_revenue > 0 else 0.0
        },
        "products": report
    }

# --- FACTURACIÓN ELECTRÓNICA ENDPOINTS ---

@router.get("/billing/profiles", response_model=List[schemas.PerfilFacturacionResponse])
def get_billing_profiles(q: Optional[str] = None, db: Session = Depends(get_db), current_user: models.Usuario = Depends(get_current_user)):
    query = db.query(models.PerfilFacturacion).filter(models.PerfilFacturacion.inquilino_id == current_user.inquilino_id)
    if q:
        query = query.filter(
            (models.PerfilFacturacion.rfc.ilike(f"%{q}%")) |
            (models.PerfilFacturacion.razon_social.ilike(f"%{q}%"))
        )
    return query.all()

@router.post("/billing/profiles", response_model=schemas.PerfilFacturacionResponse)
def create_billing_profile(profile: schemas.PerfilFacturacionCreate, db: Session = Depends(get_db), current_user: models.Usuario = Depends(get_current_user)):
    existing = db.query(models.PerfilFacturacion).filter(
        models.PerfilFacturacion.rfc == profile.rfc.upper().strip(),
        models.PerfilFacturacion.inquilino_id == current_user.inquilino_id
    ).first()
    if existing:
        raise HTTPException(status_code=400, detail="Ya existe un perfil de facturación con este RFC")
        
    db_profile = models.PerfilFacturacion(
        inquilino_id=current_user.inquilino_id,
        rfc=profile.rfc.upper().strip(),
        razon_social=profile.razon_social.upper().strip(),
        regimen_fiscal=profile.regimen_fiscal,
        codigo_postal=profile.codigo_postal,
        correo=profile.correo.lower().strip()
    )
    db.add(db_profile)
    db.commit()
    db.refresh(db_profile)
    return db_profile

@router.put("/billing/profiles/{profile_id}", response_model=schemas.PerfilFacturacionResponse)
def update_billing_profile(profile_id: int, profile: schemas.PerfilFacturacionCreate, db: Session = Depends(get_db), current_user: models.Usuario = Depends(get_current_user)):
    db_profile = db.query(models.PerfilFacturacion).filter(
        models.PerfilFacturacion.id == profile_id,
        models.PerfilFacturacion.inquilino_id == current_user.inquilino_id
    ).first()
    if not db_profile:
        raise HTTPException(status_code=404, detail="Perfil de facturación no encontrado")
        
    db_profile.rfc = profile.rfc.upper().strip()
    db_profile.razon_social = profile.razon_social.upper().strip()
    db_profile.regimen_fiscal = profile.regimen_fiscal
    db_profile.codigo_postal = profile.codigo_postal
    db_profile.correo = profile.correo.lower().strip()
    
    db.commit()
    db.refresh(db_profile)
    return db_profile

@router.get("/billing/tickets/{ticket_id}")
def get_ticket_details(ticket_id: int, db: Session = Depends(get_db), current_user: models.Usuario = Depends(get_current_user)):
    # Buscamos la línea de venta de referencia
    ref_sale = db.query(models.HistorialVenta).filter(
        models.HistorialVenta.id == ticket_id,
        models.HistorialVenta.inquilino_id == current_user.inquilino_id
    ).first()
    if not ref_sale:
        raise HTTPException(status_code=404, detail="Ticket de venta no encontrado")
        
    # Agrupamos todas las ventas que comparten la misma fecha (creado_en) y cajero
    sales = db.query(
        models.HistorialVenta.id,
        models.HistorialVenta.producto_id,
        models.HistorialVenta.variante_id,
        models.HistorialVenta.cantidad,
        models.HistorialVenta.price_sold,
        models.HistorialVenta.discount,
        models.HistorialVenta.creado_en,
        models.HistorialVenta.payment_method,
        models.HistorialVenta.factura_id,
        models.HistorialVenta.cancelado,
        models.Producto.nombre.label("nombre_producto"),
        models.Producto.clave_sat.label("product_sat_key"),
        models.Producto.clave_unidad_sat.label("product_sat_unit_key")
    ).join(
        models.Producto, models.Producto.id == models.HistorialVenta.producto_id
    ).filter(
        models.HistorialVenta.creado_en == ref_sale.creado_en,
        models.HistorialVenta.inquilino_id == current_user.inquilino_id
    ).all()
    
    if not sales:
        raise HTTPException(status_code=404, detail="No se encontraron artículos para este ticket")
        
    elementos = []
    subtotal = 0.0
    discount_total = 0.0
    
    for s in sales:
        precio = s.price_sold or 0.0
        item_subtotal = precio * s.cantidad
        subtotal += item_subtotal
        discount_total += s.discount or 0.0
        
        elementos.append({
            "venta_id": s.id,
            "producto_id": s.producto_id,
            "nombre_producto": s.nombre_producto,
            "cantidad": s.cantidad,
            "precio": precio,
            "discount": s.discount,
            "total": item_subtotal - s.discount,
            "clave_sat": s.product_sat_key,
            "clave_unidad_sat": s.product_sat_unit_key,
            "cancelado": s.cancelado
        })
        
    settings = db.query(models.ConfiguracionesTienda).filter(models.ConfiguracionesTienda.id == 1).first()
    if not settings:
        settings = models.ConfiguracionesTienda(
            id=1,
            nombre_tienda="ABARROTES ED & E",
            rfc="AED180425EE3",
            telefono="8112345678",
            correo="ventas@abarrotesede.com",
            direccion="Av. Constitución #450, Monterrey, N.L. C.P. 64000",
            tasa_impuesto=16.0,
            pie_ticket="¡Gracias por su compra!"
        )
        db.add(settings)
        db.commit()
        db.refresh(settings)

    tax_factor = 1 + (settings.tasa_impuesto / 100)
        
    taxes_total = round((subtotal - discount_total) - ((subtotal - discount_total) / tax_factor), 2)
    total = round(subtotal - discount_total, 2)
    
    # Comprobar si ya está facturada
    factura_id = sales[0].factura_id
    invoice = None
    if factura_id:
        inv_record = db.query(models.Factura).filter(models.Factura.id == factura_id).first()
        if inv_record:
            invoice = {
                "id": inv_record.id,
                "uuid": inv_record.uuid,
                "creado_en": inv_record.creado_en,
                "estado": inv_record.estado
            }
            
    return {
        "ticket_id": ticket_id,
        "creado_en": ref_sale.creado_en,
        "payment_method": ref_sale.payment_method,
        "elementos": elementos,
        "subtotal": round((subtotal - discount_total) / tax_factor, 2),
        "discount": discount_total,
        "taxes": taxes_total,
        "total": total,
        "invoice": invoice,
        "cancelado": any(s.cancelado for s in sales)
    }

@router.post("/billing/invoice", response_model=schemas.FacturaResponse)
def create_invoice(req: schemas.FacturaPeticionCreate, db: Session = Depends(get_db), current_user: models.Usuario = Depends(get_current_user)):
    # 1. Obtener los registros de venta (sale_history) que se van a facturar
    sale_ids = []
    if req.venta_id:
        sale_ids.append(req.venta_id)
    if req.sale_ids:
        sale_ids.extend(req.sale_ids)
        
    if not sale_ids:
        raise HTTPException(status_code=400, detail="Debe especificar al menos un ID de venta para facturar")
        
    # Obtener las ventas físicas de la BD
    sales = db.query(models.HistorialVenta).filter(models.HistorialVenta.id.in_(sale_ids)).all()
    if not sales:
        raise HTTPException(status_code=404, detail="No se encontraron registros de venta para facturar")
        
    # Validar que ninguna venta esté ya facturada
    for s in sales:
        if s.factura_id:
            raise HTTPException(status_code=400, detail=f"La venta con ID {s.id} ya se encuentra asociada a una factura activa.")
        if s.cancelado:
            raise HTTPException(status_code=400, detail=f"La venta con ID {s.id} está cancelada y no se puede facturar.")
            
    # 2. Resolver Perfil Fiscal
    billing_profile = None
    if req.billing_profile_id:
        profile_record = db.query(models.PerfilFacturacion).filter(models.PerfilFacturacion.id == req.billing_profile_id).first()
        if not profile_record:
            raise HTTPException(status_code=404, detail="Perfil de facturación no encontrado")
        billing_profile = {
            "rfc": profile_record.rfc,
            "razon_social": profile_record.razon_social,
            "regimen_fiscal": profile_record.regimen_fiscal,
            "codigo_postal": profile_record.codigo_postal,
            "correo": profile_record.correo
        }
    elif req.new_billing_profile:
        # Registrar perfil al vuelo
        profile_data = req.new_billing_profile
        existing = db.query(models.PerfilFacturacion).filter(models.PerfilFacturacion.rfc == profile_data.rfc.upper().strip()).first()
        if existing:
            # Reutilizar existente
            profile_record = existing
        else:
            profile_record = models.PerfilFacturacion(
                rfc=profile_data.rfc.upper().strip(),
                razon_social=profile_data.razon_social.upper().strip(),
                regimen_fiscal=profile_data.regimen_fiscal,
                codigo_postal=profile_data.codigo_postal,
                correo=profile_data.correo.lower().strip()
            )
            db.add(profile_record)
            db.commit()
            db.refresh(profile_record)
            
        billing_profile = {
            "rfc": profile_record.rfc,
            "razon_social": profile_record.razon_social,
            "regimen_fiscal": profile_record.regimen_fiscal,
            "codigo_postal": profile_record.codigo_postal,
            "correo": profile_record.correo
        }
    else:
        # Asumir Público en General
        billing_profile = {
            "rfc": "XAXX010101000",
            "razon_social": "PÚBLICO EN GENERAL",
            "regimen_fiscal": "616",  # Sin obligaciones fiscales
            "codigo_postal": "64000",
            "correo": "ventas@abarrotesede.com"
        }
        
    # 3. Preparar lista de conceptos para el generador
    elementos = []
    subtotal = 0.0
    discount_total = 0.0
    
    for s in sales:
        prod = db.query(models.Producto).filter(models.Producto.id == s.producto_id).first()
        name = prod.nombre if prod else "PRODUCTO DESCONOCIDO"
        clave_sat = (prod.clave_sat if prod else "01010101") or "01010101"
        clave_unidad_sat = (prod.clave_unidad_sat if prod else "H87") or "H87"
        
        precio = s.price_sold or 0.0
        subtotal += precio * s.cantidad
        discount_total += s.discount or 0.0
        
        elementos.append({
            "nombre": name,
            "cantidad": s.cantidad,
            "precio": precio,
            "clave_sat": clave_sat,
            "clave_unidad_sat": clave_unidad_sat
        })
        
    total = round(subtotal - discount_total, 2)
    
    # 4. Generar metadatos del Timbre
    invoice_uuid = str(uuid.uuid4()).upper()
    timestamp_str = datetime.utcnow().strftime('%Y-%m-%dT%H:%M:%S')
    
    # Directorio para almacenar los archivos locales
    invoices_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "static", "invoices")
    os.makedirs(invoices_dir, exist_ok=True)
    
    xml_path = os.path.join(invoices_dir, f"{invoice_uuid}.xml")
    pdf_path = os.path.join(invoices_dir, f"{invoice_uuid}.pdf")
    
    # Generar archivos físicos con configuración dinámica
    settings = db.query(models.ConfiguracionesTienda).filter(models.ConfiguracionesTienda.id == 1).first()
    store_settings_dict = None
    if settings:
        store_settings_dict = {
            "nombre_tienda": settings.nombre_tienda,
            "rfc": settings.rfc,
            "telefono": settings.telefono,
            "correo": settings.correo,
            "direccion": settings.direccion,
            "tasa_impuesto": settings.tasa_impuesto,
            "pie_ticket": settings.pie_ticket
        }

    try:
        # Generar XML
        xml_content = facturacion.generate_cfdi_xml(
            sale_info={"payment_method": sales[0].payment_method, "discount": discount_total},
            billing_profile=billing_profile,
            elementos=elementos,
            invoice_uuid=invoice_uuid,
            timestamp_str=timestamp_str,
            store_settings=store_settings_dict
        )
        with open(xml_path, "w", encoding="utf-8") as f:
            f.write(xml_content)
            
        # Generar PDF
        facturacion.generate_cfdi_pdf(
            pdf_path=pdf_path,
            sale_info={"payment_method": sales[0].payment_method, "discount": discount_total},
            billing_profile=billing_profile,
            elementos=elementos,
            invoice_uuid=invoice_uuid,
            timestamp_str=timestamp_str,
            store_settings=store_settings_dict
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error al generar los documentos fiscales: {str(e)}")
        
    # 5. Guardar en Base de Datos
    db_invoice = models.Factura(
        inquilino_id=current_user.inquilino_id,
        uuid=invoice_uuid,
        monto_total=total,
        xml_url=f"/api/billing/invoices/{invoice_uuid}/xml",
        pdf_url=f"/api/billing/invoices/{invoice_uuid}/pdf",
        creado_en=timestamp_str,
        estado="active"
    )
    db.add(db_invoice)
    db.commit()
    db.refresh(db_invoice)
    
    # Vincular los SaleHistory
    for s in sales:
        s.factura_id = db_invoice.id
    db.commit()
    
    return db_invoice

@router.get("/billing/invoices", response_model=List[schemas.FacturaResponse])
def get_invoices(db: Session = Depends(get_db), current_user: models.Usuario = Depends(get_current_user)):
    return db.query(models.Factura).filter(models.Factura.inquilino_id == current_user.inquilino_id).order_by(models.Factura.id.desc()).all()

@router.get("/billing/invoices/{invoice_uuid}/xml")
def download_invoice_xml(invoice_uuid: str, db: Session = Depends(get_db)):
    invoice = db.query(models.Factura).filter(models.Factura.uuid == invoice_uuid).first()
    if not invoice:
        raise HTTPException(status_code=404, detail="Factura no encontrada")
        
    invoices_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "static", "invoices")
    file_path = os.path.join(invoices_dir, f"{invoice_uuid}.xml")
    if not os.path.exists(file_path):
        raise HTTPException(status_code=404, detail="El archivo XML de la factura no existe en el servidor")
        
    return FileResponse(file_path, media_type="application/xml", filename=f"CFDI_{invoice_uuid}.xml")

@router.get("/billing/invoices/{invoice_uuid}/pdf")
def download_invoice_pdf(invoice_uuid: str, db: Session = Depends(get_db)):
    invoice = db.query(models.Factura).filter(models.Factura.uuid == invoice_uuid).first()
    if not invoice:
        raise HTTPException(status_code=404, detail="Factura no encontrada")
        
    invoices_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "static", "invoices")
    file_path = os.path.join(invoices_dir, f"{invoice_uuid}.pdf")
    if not os.path.exists(file_path):
        raise HTTPException(status_code=404, detail="El archivo PDF de la factura no existe en el servidor")
        
    return FileResponse(file_path, media_type="application/pdf", filename=f"Factura_{invoice_uuid}.pdf")

@router.post("/billing/invoices/{factura_id}/cancel", response_model=schemas.FacturaResponse)
def cancel_invoice(factura_id: int, db: Session = Depends(get_db), current_user: models.Usuario = Depends(get_current_user)):
    if current_user.rol not in ["admin", "supervisor"]:
        raise HTTPException(status_code=403, detail="No tienes permisos suficientes para cancelar facturas")
        
    invoice = db.query(models.Factura).filter(
        models.Factura.id == factura_id,
        models.Factura.inquilino_id == current_user.inquilino_id
    ).first()
    if not invoice:
        raise HTTPException(status_code=404, detail="Factura no encontrada")
        
    if invoice.estado == "cancelled":
        raise HTTPException(status_code=400, detail="Esta factura ya se encuentra cancelada")
        
    # Cambiar estatus de la factura
    invoice.estado = "cancelled"
    
    # Desvincular ventas asociadas (o dejarlas marcadas, pero para permitir volver a facturarlas, ponemos su factura_id en NULL!)
    db.query(models.HistorialVenta).filter(
        models.HistorialVenta.factura_id == invoice.id,
        models.HistorialVenta.inquilino_id == current_user.inquilino_id
    ).update({models.HistorialVenta.factura_id: None})
    
    db.commit()
    db.refresh(invoice)
    return invoice


# --- PROVEEDORES (SUPPLIERS) ENDPOINTS ---

@router.get("/suppliers/", response_model=List[schemas.ProveedorResponse])
def get_suppliers(db: Session = Depends(get_db), current_user: models.Usuario = Depends(get_current_user)):
    if current_user.rol not in ['admin', 'supervisor']:
        raise HTTPException(status_code=403, detail="No tienes permisos suficientes")
    return db.query(models.Proveedor).filter(models.Proveedor.inquilino_id == current_user.inquilino_id).order_by(models.Proveedor.nombre).all()

@router.post("/suppliers/", response_model=schemas.ProveedorResponse)
def create_supplier(proveedor: schemas.ProveedorCreate, db: Session = Depends(get_db), current_user: models.Usuario = Depends(get_current_user)):
    if current_user.rol not in ['admin', 'supervisor']:
        raise HTTPException(status_code=403, detail="No tienes permisos suficientes")
    
    if proveedor.rfc and proveedor.rfc.strip():
        existing = db.query(models.Proveedor).filter(
            models.Proveedor.rfc == proveedor.rfc.upper().strip(),
            models.Proveedor.inquilino_id == current_user.inquilino_id
        ).first()
        if existing:
            raise HTTPException(status_code=400, detail="Ya existe un proveedor con ese RFC")

    db_supplier = models.Proveedor(
        inquilino_id=current_user.inquilino_id,
        name=proveedor.nombre.strip(),
        rfc=proveedor.rfc.upper().strip() if proveedor.rfc else None,
        telefono=proveedor.telefono.strip() if proveedor.telefono else None,
        correo=proveedor.correo.lower().strip() if proveedor.correo else None,
        direccion=proveedor.direccion.strip() if proveedor.direccion else None,
        notas=proveedor.notas.strip() if proveedor.notas else None
    )
    db.add(db_supplier)
    db.commit()
    db.refresh(db_supplier)
    return db_supplier

@router.put("/suppliers/{proveedor_id}", response_model=schemas.ProveedorResponse)
def update_supplier(proveedor_id: int, proveedor: schemas.ProveedorCreate, db: Session = Depends(get_db), current_user: models.Usuario = Depends(get_current_user)):
    if current_user.rol not in ['admin', 'supervisor']:
        raise HTTPException(status_code=403, detail="No tienes permisos suficientes")
    
    db_supplier = db.query(models.Proveedor).filter(
        models.Proveedor.id == proveedor_id,
        models.Proveedor.inquilino_id == current_user.inquilino_id
    ).first()
    if not db_supplier:
        raise HTTPException(status_code=404, detail="Proveedor no encontrado")

    if proveedor.rfc and proveedor.rfc.strip():
        existing = db.query(models.Proveedor).filter(
            models.Proveedor.rfc == proveedor.rfc.upper().strip(), 
            models.Proveedor.id != proveedor_id,
            models.Proveedor.inquilino_id == current_user.inquilino_id
        ).first()
        if existing:
            raise HTTPException(status_code=400, detail="Ya existe otro proveedor con ese RFC")

    db_supplier.nombre = proveedor.nombre.strip()
    db_supplier.rfc = proveedor.rfc.upper().strip() if proveedor.rfc else None
    db_supplier.telefono = proveedor.telefono.strip() if proveedor.telefono else None
    db_supplier.correo = proveedor.correo.lower().strip() if proveedor.correo else None
    db_supplier.direccion = proveedor.direccion.strip() if proveedor.direccion else None
    db_supplier.notas = proveedor.notas.strip() if proveedor.notas else None

    db.commit()
    db.refresh(db_supplier)
    return db_supplier

@router.delete("/suppliers/{proveedor_id}")
def delete_supplier(proveedor_id: int, db: Session = Depends(get_db), current_user: models.Usuario = Depends(get_current_user)):
    if current_user.rol not in ['admin', 'supervisor']:
        raise HTTPException(status_code=403, detail="No tienes permisos suficientes")
    
    db_supplier = db.query(models.Proveedor).filter(
        models.Proveedor.id == proveedor_id,
        models.Proveedor.inquilino_id == current_user.inquilino_id
    ).first()
    if not db_supplier:
        raise HTTPException(status_code=404, detail="Proveedor no encontrado")
    
    associated_purchases = db.query(models.Compra).filter(
        models.Compra.proveedor_id == proveedor_id,
        models.Compra.inquilino_id == current_user.inquilino_id
    ).first()
    if associated_purchases:
        raise HTTPException(
            status_code=400, 
            detail="No se puede eliminar el proveedor porque tiene compras/entradas asociadas. Considere editarlo o dejarlo inactivo."
        )

    db.delete(db_supplier)
    db.commit()
    return {"message": "Proveedor eliminado exitosamente"}


# --- COMPRAS / ENTRADAS (PURCHASES) ENDPOINTS ---

@router.get("/purchases/", response_model=List[schemas.CompraResponse])
def get_purchases(db: Session = Depends(get_db), current_user: models.Usuario = Depends(get_current_user)):
    if current_user.rol not in ['admin', 'supervisor']:
        raise HTTPException(status_code=403, detail="No tienes permisos suficientes")
        
    purchases = db.query(models.Compra).filter(models.Compra.inquilino_id == current_user.inquilino_id).order_by(models.Compra.id.desc()).all()
    
    result = []
    for p in purchases:
        supplier_name = "Compra Directa / Sin Proveedor"
        if p.proveedor_id:
            proveedor = db.query(models.Proveedor).filter(
                models.Proveedor.id == p.proveedor_id,
                models.Proveedor.inquilino_id == current_user.inquilino_id
            ).first()
            if proveedor:
                supplier_name = proveedor.nombre
                
        user_name = "Desconocido"
        if p.usuario_id:
            user = db.query(models.Usuario).filter(
                models.Usuario.id == p.usuario_id,
                models.Usuario.inquilino_id == current_user.inquilino_id
            ).first()
            if user:
                user_name = user.nombre_completo
                
        items_res = []
        for item in p.items:
            prod = db.query(models.Producto).filter(
                models.Producto.id == item.producto_id,
                models.Producto.inquilino_id == current_user.inquilino_id
            ).first()
            prod_name = prod.nombre if prod else f"Producto ID {item.producto_id}"
            if item.variante_id:
                var = db.query(models.VarianteProducto).filter(
                    models.VarianteProducto.id == item.variante_id,
                    models.VarianteProducto.inquilino_id == current_user.inquilino_id
                ).first()
                if var:
                    prod_name += f" ({var.nombre})"
            
            items_res.append(
                schemas.ElementoCompraResponse(
                    id=item.id,
                    compra_id=item.compra_id,
                    producto_id=item.producto_id,
                    variante_id=item.variante_id,
                    cantidad=item.cantidad,
                    precio_costo=item.precio_costo,
                    precio=item.precio,
                    nombre_producto=prod_name
                )
            )
            
        result.append(
            schemas.CompraResponse(
                id=p.id,
                proveedor_id=p.proveedor_id,
                numero_factura=p.numero_factura,
                costo_total=p.costo_total,
                creado_en=p.creado_en,
                notas=p.notas,
                usuario_id=p.usuario_id,
                elementos=items_res,
                nombre_proveedor=supplier_name,
                nombre_usuario=user_name
            )
        )
    return result

@router.get("/purchases/{compra_id}", response_model=schemas.CompraResponse)
def get_purchase_details(compra_id: int, db: Session = Depends(get_db), current_user: models.Usuario = Depends(get_current_user)):
    if current_user.rol not in ['admin', 'supervisor']:
        raise HTTPException(status_code=403, detail="No tienes permisos suficientes")
        
    p = db.query(models.Compra).filter(
        models.Compra.id == compra_id,
        models.Compra.inquilino_id == current_user.inquilino_id
    ).first()
    if not p:
        raise HTTPException(status_code=404, detail="Compra no encontrada")
        
    supplier_name = "Compra Directa / Sin Proveedor"
    if p.proveedor_id:
        proveedor = db.query(models.Proveedor).filter(
            models.Proveedor.id == p.proveedor_id,
            models.Proveedor.inquilino_id == current_user.inquilino_id
        ).first()
        if proveedor:
            supplier_name = proveedor.nombre
            
    user_name = "Desconocido"
    if p.usuario_id:
        user = db.query(models.Usuario).filter(
            models.Usuario.id == p.usuario_id,
            models.Usuario.inquilino_id == current_user.inquilino_id
        ).first()
        if user:
            user_name = user.nombre_completo
            
    items_res = []
    for item in p.items:
        prod = db.query(models.Producto).filter(
            models.Producto.id == item.producto_id,
            models.Producto.inquilino_id == current_user.inquilino_id
        ).first()
        prod_name = prod.nombre if prod else f"Producto ID {item.producto_id}"
        if item.variante_id:
            var = db.query(models.VarianteProducto).filter(
                models.VarianteProducto.id == item.variante_id,
                models.VarianteProducto.inquilino_id == current_user.inquilino_id
            ).first()
            if var:
                prod_name += f" ({var.nombre})"
        
        items_res.append(
            schemas.ElementoCompraResponse(
                id=item.id,
                compra_id=item.compra_id,
                producto_id=item.producto_id,
                variante_id=item.variante_id,
                cantidad=item.cantidad,
                precio_costo=item.precio_costo,
                precio=item.precio,
                nombre_producto=prod_name
            )
        )
        
    return schemas.CompraResponse(
        id=p.id,
        proveedor_id=p.proveedor_id,
        numero_factura=p.numero_factura,
        costo_total=p.costo_total,
        creado_en=p.creado_en,
        notas=p.notas,
        usuario_id=p.usuario_id,
        elementos=items_res,
        nombre_proveedor=supplier_name,
        nombre_usuario=user_name
    )

@router.post("/purchases/", response_model=schemas.CompraResponse)
def create_purchase(purchase_data: schemas.CompraCreate, db: Session = Depends(get_db), current_user: models.Usuario = Depends(get_current_user)):
    if current_user.rol not in ['admin', 'supervisor']:
        raise HTTPException(status_code=403, detail="No tienes permisos suficientes")
        
    if not purchase_data.items:
        raise HTTPException(status_code=400, detail="Debe añadir al menos un artículo a la compra.")

    try:
        now_str = datetime.utcnow().isoformat()
        
        db_purchase = models.Compra(
            inquilino_id=current_user.inquilino_id,
            proveedor_id=purchase_data.proveedor_id,
            numero_factura=purchase_data.numero_factura.strip() if purchase_data.numero_factura else None,
            costo_total=0.0,
            creado_en=now_str,
            notas=purchase_data.notas.strip() if purchase_data.notas else None,
            usuario_id=current_user.id
        )
        db.add(db_purchase)
        db.commit()
        db.refresh(db_purchase)
        
        costo_total = 0.0
        
        for item in purchase_data.items:
            producto = db.query(models.Producto).filter(
                models.Producto.id == item.producto_id,
                models.Producto.inquilino_id == current_user.inquilino_id
            ).with_for_update().first()
            if not producto:
                raise HTTPException(status_code=404, detail=f"El producto con ID {item.producto_id} no existe.")
            
            if item.variante_id:
                variante = db.query(models.VarianteProducto).filter(
                    models.VarianteProducto.id == item.variante_id,
                    models.VarianteProducto.producto_id == item.producto_id,
                    models.VarianteProducto.inquilino_id == current_user.inquilino_id
                ).with_for_update().first()
                if not variante:
                    raise HTTPException(
                        status_code=404, 
                        detail=f"La variante ID {item.variante_id} para el producto ID {item.producto_id} no existe."
                    )
                
                variante.cantidad += item.cantidad
                variante.precio_costo = item.precio_costo
                if item.precio is not None and item.precio > 0:
                    variante.precio = item.precio
                
                producto.cantidad += item.cantidad
            else:
                producto.cantidad += item.cantidad
                producto.precio_costo = item.precio_costo
                if item.precio is not None and item.precio > 0:
                    producto.precio = item.precio
                    
            item_cost = item.precio_costo * item.cantidad
            costo_total += item_cost
            
            db_item = models.ElementoCompra(
                compra_id=db_purchase.id,
                producto_id=item.producto_id,
                variante_id=item.variante_id,
                cantidad=item.cantidad,
                precio_costo=item.precio_costo,
                precio=item.precio
            )
            db.add(db_item)
            
        db_purchase.costo_total = costo_total
        db.commit()
        db.refresh(db_purchase)
        
        return get_purchase_details(compra_id=db_purchase.id, db=db, current_user=current_user)
        
    except Exception as e:
        db.rollback()
        if isinstance(e, HTTPException):
            raise e
        raise HTTPException(status_code=400, detail=f"Error al procesar la compra: {str(e)}")


# --- AJUSTES DE TIENDA (STORE SETTINGS) ENDPOINTS ---

@router.get("/settings", response_model=schemas.ConfiguracionesTiendaResponse)
def get_store_settings(db: Session = Depends(get_db), current_user: models.Usuario = Depends(get_current_user)):
    settings = db.query(models.ConfiguracionesTienda).filter(models.ConfiguracionesTienda.inquilino_id == current_user.inquilino_id).first()
    if not settings:
        settings = models.ConfiguracionesTienda(
            inquilino_id=current_user.inquilino_id,
            nombre_tienda="ABARROTES ED & E",
            rfc="AED180425EE3",
            telefono="8112345678",
            correo="ventas@abarrotesede.com",
            direccion="Av. Constitución #450, Monterrey, N.L. C.P. 64000",
            tasa_impuesto=16.0,
            pie_ticket="¡Gracias por su compra!"
        )
        db.add(settings)
        db.commit()
        db.refresh(settings)
    return settings

@router.put("/settings", response_model=schemas.ConfiguracionesTiendaResponse)
def update_store_settings(settings_data: schemas.ConfiguracionesTiendaCreate, db: Session = Depends(get_db), current_user: models.Usuario = Depends(get_current_user)):
    if current_user.rol not in ['admin', 'supervisor']:
        raise HTTPException(status_code=403, detail="No tienes permisos suficientes para cambiar la configuración.")
    
    settings = db.query(models.ConfiguracionesTienda).filter(models.ConfiguracionesTienda.inquilino_id == current_user.inquilino_id).first()
    if not settings:
        settings = models.ConfiguracionesTienda(inquilino_id=current_user.inquilino_id)
        db.add(settings)
        
    for key, value in settings_data.dict().items():
        setattr(settings, key, value)
        
    db.commit()
    db.refresh(settings)
    return settings


# --- CLIENTES Y CRÉDITOS (CUSTOMERS & CREDIT) ENDPOINTS ---

@router.get("/customers", response_model=List[schemas.ClienteResponse])
def get_customers(q: Optional[str] = None, db: Session = Depends(get_db), current_user: models.Usuario = Depends(get_current_user)):
    query = db.query(models.Cliente).filter(models.Cliente.inquilino_id == current_user.inquilino_id)
    if q:
        query = query.filter(
            (models.Cliente.nombre.ilike(f"%{q}%")) |
            (models.Cliente.telefono.ilike(f"%{q}%"))
        )
    return query.order_by(models.Cliente.nombre).all()

@router.post("/customers", response_model=schemas.ClienteResponse)
def create_customer(customer_data: schemas.ClienteCreate, db: Session = Depends(get_db), current_user: models.Usuario = Depends(get_current_user)):
    if current_user.rol not in ['admin', 'supervisor']:
        raise HTTPException(status_code=403, detail="No tienes permisos suficientes para registrar clientes.")
        
    existing = db.query(models.Cliente).filter(
        models.Cliente.nombre.ilike(customer_data.nombre.strip()),
        models.Cliente.inquilino_id == current_user.inquilino_id
    ).first()
    if existing:
        raise HTTPException(status_code=400, detail="Ya existe un cliente con ese nombre.")
        
    db_customer = models.Cliente(
        inquilino_id=current_user.inquilino_id,
        name=customer_data.nombre.strip(),
        telefono=customer_data.telefono.strip() if customer_data.telefono else None,
        correo=customer_data.correo.lower().strip() if customer_data.correo else None,
        limite_credito=customer_data.limite_credito,
        saldo_actual=0.0
    )
    db.add(db_customer)
    db.commit()
    db.refresh(db_customer)
    return db_customer

@router.put("/customers/{cliente_id}", response_model=schemas.ClienteResponse)
def update_customer(cliente_id: int, customer_data: schemas.ClienteCreate, db: Session = Depends(get_db), current_user: models.Usuario = Depends(get_current_user)):
    if current_user.rol not in ['admin', 'supervisor']:
        raise HTTPException(status_code=403, detail="No tienes permisos suficientes para modificar clientes.")

    db_customer = db.query(models.Cliente).filter(
        models.Cliente.id == cliente_id,
        models.Cliente.inquilino_id == current_user.inquilino_id
    ).first()
    if not db_customer:
        raise HTTPException(status_code=404, detail="Cliente no encontrado.")
        
    existing = db.query(models.Cliente).filter(
        models.Cliente.nombre.ilike(customer_data.nombre.strip()), 
        models.Cliente.id != cliente_id,
        models.Cliente.inquilino_id == current_user.inquilino_id
    ).first()
    if existing:
        raise HTTPException(status_code=400, detail="Ya existe otro cliente con ese nombre.")
        
    db_customer.nombre = customer_data.nombre.strip()
    db_customer.telefono = customer_data.telefono.strip() if customer_data.telefono else None
    db_customer.correo = customer_data.correo.lower().strip() if customer_data.correo else None
    db_customer.limite_credito = customer_data.limite_credito
    
    db.commit()
    db.refresh(db_customer)
    return db_customer

@router.delete("/customers/{cliente_id}")
def delete_customer(cliente_id: int, db: Session = Depends(get_db), current_user: models.Usuario = Depends(get_current_user)):
    if current_user.rol not in ['admin', 'supervisor']:
        raise HTTPException(status_code=403, detail="No tienes permisos suficientes para eliminar clientes.")
        
    db_customer = db.query(models.Cliente).filter(
        models.Cliente.id == cliente_id,
        models.Cliente.inquilino_id == current_user.inquilino_id
    ).first()
    if not db_customer:
        raise HTTPException(status_code=404, detail="Cliente no encontrado.")
        
    if db_customer.saldo_actual > 0:
        raise HTTPException(status_code=400, detail="No se puede eliminar el cliente porque tiene un saldo deudor pendiente.")
        
    db.delete(db_customer)
    db.commit()
    return {"message": "Cliente eliminado exitosamente"}

@router.post("/customers/{cliente_id}/pay", response_model=schemas.PagoClienteResponse)
def register_customer_payment(cliente_id: int, payment_data: schemas.PagoClienteCreate, db: Session = Depends(get_db), current_user: models.Usuario = Depends(get_current_user)):
    shift = db.query(models.Turno).filter(
        models.Turno.usuario_id == current_user.id,
        models.Turno.inquilino_id == current_user.inquilino_id,
        models.Turno.estado == "open"
    ).first()
    if not shift and current_user.rol != 'admin':
        raise HTTPException(status_code=400, detail="Debe tener un turno de caja abierto para registrar un abono.")
        
    if current_user.rol in ["cajero", "user"]:
        auth_user = payment_data.auth_username or payment_data.usuario_autorizacion
        auth_pass = payment_data.auth_password or payment_data.contrasena_autorizacion
        
        if not auth_user or not auth_pass:
            raise HTTPException(status_code=403, detail="Para registrar abonos de clientes como cajero se requiere usuario y contraseña de Administrador o Supervisor.")
            
        sup = db.query(models.Usuario).filter(
            models.Usuario.nombre_usuario == auth_user.strip(),
            models.Usuario.inquilino_id == current_user.inquilino_id
        ).first()
        
        if not sup or not auth.verify_password(auth_pass, sup.contrasena_encriptada):
            raise HTTPException(status_code=403, detail="Usuario o contraseña del supervisor incorrectos.")
            
        if sup.rol not in ["admin", "supervisor"]:
            raise HTTPException(status_code=403, detail="El usuario autorizador no tiene permisos de Supervisor o Administrador.")

    customer = db.query(models.Cliente).filter(
        models.Cliente.id == cliente_id,
        models.Cliente.inquilino_id == current_user.inquilino_id
    ).with_for_update().first()
    if not customer:
        raise HTTPException(status_code=404, detail="Cliente no encontrado.")
        
    if payment_data.monto <= 0:
        raise HTTPException(status_code=400, detail="El monto del abono debe ser mayor a 0.")
        
    db_payment = models.PagoCliente(
        inquilino_id=current_user.inquilino_id,
        cliente_id=cliente_id,
        turno_id=shift.id if shift else None,
        usuario_id=current_user.id,
        monto=payment_data.monto,
        creado_en=datetime.utcnow().isoformat(),
        notas=payment_data.notas.strip() if payment_data.notas else None
    )
    
    customer.saldo_actual -= payment_data.monto
    
    if shift:
        db_mov = models.MovimientoCaja(
            turno_id=shift.id,
            tipo="entrada",
            monto=payment_data.monto,
            motivo=f"Abono de cliente: {customer.nombre}",
            creado_en=datetime.utcnow().isoformat()
        )
        shift.efectivo_final_esperado += payment_data.monto
        db.add(db_mov)
        
    db.add(db_payment)
    db.commit()
    db.refresh(db_payment)
    return db_payment

@router.get("/customers/{cliente_id}/history")
def get_customer_history(cliente_id: int, db: Session = Depends(get_db), current_user: models.Usuario = Depends(get_current_user)):
    customer = db.query(models.Cliente).filter(
        models.Cliente.id == cliente_id,
        models.Cliente.inquilino_id == current_user.inquilino_id
    ).first()
    if not customer:
        raise HTTPException(status_code=404, detail="Cliente no encontrado.")
        
    sales = db.query(
        models.HistorialVenta.id,
        models.HistorialVenta.creado_en,
        models.HistorialVenta.payment_method,
        models.HistorialVenta.price_sold,
        models.HistorialVenta.cantidad,
        models.HistorialVenta.discount,
        models.Producto.nombre.label("nombre_producto")
    ).join(
        models.Producto, models.Producto.id == models.HistorialVenta.producto_id
    ).filter(
        models.HistorialVenta.cliente_id == cliente_id,
        models.HistorialVenta.inquilino_id == current_user.inquilino_id
    ).all()
    
    payments = db.query(models.PagoCliente).filter(
        models.PagoCliente.cliente_id == cliente_id,
        models.PagoCliente.inquilino_id == current_user.inquilino_id
    ).all()
    
    history = []
    sales_by_ticket = {}
    for s in sales:
        total_item = (s.price_sold * s.cantidad) - s.discount
        date_key = s.creado_en
        if date_key not in sales_by_ticket:
            sales_by_ticket[date_key] = {
                "id": s.id,
                "tipo": "compra",
                "payment_method": s.payment_method,
                "creado_en": s.creado_en,
                "monto": 0.0,
                "details": []
            }
        sales_by_ticket[date_key]["monto"] += total_item
        sales_by_ticket[date_key]["details"].append(f"{s.cantidad}x {s.nombre_producto}")
        
    for ticket in sales_by_ticket.values():
        history.append({
            "id": ticket["id"],
            "tipo": "compra_credito" if ticket["payment_method"] == "credito" else "compra_asociada",
            "description": ", ".join(ticket["details"]),
            "monto": round(ticket["monto"], 2),
            "creado_en": ticket["creado_en"]
        })
        
    for p in payments:
        history.append({
            "id": p.id,
            "tipo": "abono",
            "description": p.notas or "Abono a cuenta",
            "monto": p.monto,
            "creado_en": p.creado_en
        })
        
    history = sorted(history, key=lambda x: x["creado_en"], reverse=True)
    return {
        "customer": {
            "id": customer.id,
            "nombre": customer.nombre,
            "saldo_actual": customer.saldo_actual,
            "limite_credito": customer.limite_credito
        },
        "history": history
    }


# --- BACKUP & SECURITY ENDPOINTS ---
import json
import io
from fastapi.responses import StreamingResponse
from fastapi import UploadFile, File
from sqlalchemy import text

@router.get("/backup/export")
def export_backup_database(format: str = "json", db: Session = Depends(get_db), current_user: models.Usuario = Depends(get_current_user)):
    if current_user.rol != "admin":
        raise HTTPException(status_code=403, detail="No tienes permisos para exportar respaldos.")
        
    def serialize_table(model):
        if hasattr(model, "inquilino_id"):
            rows = db.query(model).filter(model.inquilino_id == current_user.inquilino_id).all()
        elif model.__tablename__ == "purchase_items":
            rows = db.query(models.ElementoCompra).join(models.Compra).filter(models.Compra.inquilino_id == current_user.inquilino_id).all()
        else:
            rows = db.query(model).all()
            
        result = []
        for row in rows:
            row_dict = {}
            for key in row.__mapper__.columns.keys():
                val = getattr(row, key)
                if val is not None and hasattr(val, "isoformat"):
                    val = val.isoformat()
                row_dict[key] = val
            result.append(row_dict)
        return result

    if format == "json":
        backup_data = {
            "store_settings": serialize_table(models.ConfiguracionesTienda),
            "users": serialize_table(models.Usuario),
            "suppliers": serialize_table(models.Proveedor),
            "billing_profiles": serialize_table(models.PerfilFacturacion),
            "customers": serialize_table(models.Cliente),
            "products": serialize_table(models.Producto),
            "product_variants": serialize_table(models.VarianteProducto),
            "invoices": serialize_table(models.Factura),
            "shifts": serialize_table(models.Turno),
            "customer_payments": serialize_table(models.PagoCliente),
            "purchases": serialize_table(models.Compra),
            "purchase_items": serialize_table(models.ElementoCompra),
            "cash_movements": serialize_table(models.MovimientoCaja),
            "sales_history": serialize_table(models.HistorialVenta),
            "product_returns": serialize_table(models.DevolucionProducto),
            "notifications": serialize_table(models.Notificacion)
        }
        json_str = json.dumps(backup_data, indent=2, ensure_ascii=False)
        stream = io.BytesIO(json_str.encode('utf-8'))
        
        headers = {
            'Content-Disposition': f'attachment; filename="tienda_backup_{datetime.now().strftime("%Y-%m-%d")}.json"'
        }
        return StreamingResponse(stream, media_type="application/json", headers=headers)
        
    elif format == "sql":
        sql_lines = []
        sql_lines.append("-- TIENDA DATABASE BACKUP SQL DUMP")
        sql_lines.append(f"-- Generated: {datetime.now().isoformat()}")
        sql_lines.append(f"-- Tenant: {current_user.inquilino_id}")
        sql_lines.append("")
        
        tables = [
            ("store_settings", models.ConfiguracionesTienda),
            ("users", models.Usuario),
            ("suppliers", models.Proveedor),
            ("billing_profiles", models.PerfilFacturacion),
            ("customers", models.Cliente),
            ("products", models.Producto),
            ("product_variants", models.VarianteProducto),
            ("invoices", models.Factura),
            ("shifts", models.Turno),
            ("customer_payments", models.PagoCliente),
            ("purchases", models.Compra),
            ("purchase_items", models.ElementoCompra),
            ("cash_movements", models.MovimientoCaja),
            ("sales_history", models.HistorialVenta),
            ("product_returns", models.DevolucionProducto),
            ("notifications", models.Notificacion)
        ]
        
        for table_name, model in tables:
            if hasattr(model, "inquilino_id"):
                rows = db.query(model).filter(model.inquilino_id == current_user.inquilino_id).all()
            elif model.__tablename__ == "purchase_items":
                rows = db.query(models.ElementoCompra).join(models.Compra).filter(models.Compra.inquilino_id == current_user.inquilino_id).all()
            else:
                rows = db.query(model).all()
                
            if not rows:
                continue
            sql_lines.append(f"\n-- Data for table {table_name}")
            
            columns = model.__mapper__.columns.keys()
            cols_str = ", ".join(columns)
            
            for row in rows:
                vals = []
                for col in columns:
                    val = getattr(row, col)
                    if val is None:
                        vals.append("NULL")
                    elif isinstance(val, (int, float)):
                        vals.append(str(val))
                    elif isinstance(val, bool):
                        vals.append("TRUE" if val else "FALSE")
                    else:
                        escaped_str = str(val).replace("'", "''")
                        vals.append(f"'{escaped_str}'")
                vals_str = ", ".join(vals)
                sql_lines.append(f"INSERT INTO {table_name} ({cols_str}) VALUES ({vals_str});")
            
        sql_str = "\n".join(sql_lines)
        stream = io.BytesIO(sql_str.encode('utf-8'))
        headers = {
            'Content-Disposition': f'attachment; filename="tienda_backup_{datetime.now().strftime("%Y-%m-%d")}.sql"'
        }
        return StreamingResponse(stream, media_type="application/sql", headers=headers)
        
    elif format == "excel":
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        
        wb = openpyxl.Workbook()
        default_sheet = wb.active
        wb.remove(default_sheet)
        
        tables = [
            ("Ajustes", models.ConfiguracionesTienda),
            ("Usuarios", models.Usuario),
            ("Proveedores", models.Proveedor),
            ("Perfiles_Facturacion", models.PerfilFacturacion),
            ("Clientes", models.Cliente),
            ("Productos", models.Producto),
            ("Variantes_Producto", models.VarianteProducto),
            ("Facturas", models.Factura),
            ("Turnos_Caja", models.Turno),
            ("Abonos_Clientes", models.PagoCliente),
            ("Compras", models.Compra),
            ("Items_Compra", models.ElementoCompra),
            ("Movimientos_Caja", models.MovimientoCaja),
            ("Historial_Ventas", models.HistorialVenta),
            ("Devoluciones", models.DevolucionProducto)
        ]
        
        title_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        cell_font = Font(name="Calibri", size=10)
        header_fill = PatternFill(start_color="064E3B", end_color="064E3B", fill_type="solid")
        align_center = Alignment(horizontal="center", vertical="center")
        align_left = Alignment(horizontal="left", vertical="center")
        
        thin_side = Side(border_style="thin", color="D1D5DB")
        thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
        
        for sheet_name, model in tables:
            ws = wb.create_sheet(title=sheet_name)
            if hasattr(model, "inquilino_id"):
                rows = db.query(model).filter(model.inquilino_id == current_user.inquilino_id).all()
            elif model.__tablename__ == "purchase_items":
                rows = db.query(models.ElementoCompra).join(models.Compra).filter(models.Compra.inquilino_id == current_user.inquilino_id).all()
            else:
                rows = db.query(model).all()
                
            columns = model.__mapper__.columns.keys()
            
            for col_idx, col_name in enumerate(columns, 1):
                cell = ws.cell(row=1, column=col_idx, value=col_name.upper())
                cell.font = title_font
                cell.fill = header_fill
                cell.alignment = align_center
                cell.border = thin_border
                
            for row_idx, row in enumerate(rows, 2):
                for col_idx, col_name in enumerate(columns, 1):
                    val = getattr(row, col_name)
                    if isinstance(val, bool):
                        val = "SÍ" if val else "NO"
                    cell = ws.cell(row=row_idx, column=col_idx, value=val)
                    cell.font = cell_font
                    cell.border = thin_border
                    if isinstance(val, (int, float)):
                        cell.alignment = Alignment(horizontal="right", vertical="center")
                    else:
                        cell.alignment = align_left
                        
            for col in ws.columns:
                max_len = max(len(str(cell.value or '')) for cell in col)
                col_letter = openpyxl.utils.get_column_letter(col[0].column)
                ws.column_dimensions[col_letter].width = max(max_len + 3, 10)
                
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        headers = {
            'Content-Disposition': f'attachment; filename="tienda_backup_{datetime.now().strftime("%Y-%m-%d")}.xlsx"'
        }
        return StreamingResponse(output, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers=headers)
        
    else:
        raise HTTPException(status_code=400, detail="Formato de respaldo no soportado.")


@router.post("/backup/import")
def import_backup_database(file: UploadFile = File(...), db: Session = Depends(get_db), current_user: models.Usuario = Depends(get_current_user)):
    if current_user.rol != "admin":
        raise HTTPException(status_code=403, detail="No tienes permisos para restaurar respaldos.")
        
    try:
        content = file.file.read()
        data = json.loads(content.decode('utf-8'))
    except Exception as e:
        raise HTTPException(status_code=400, detail="El archivo no es un JSON válido o tiene codificación incorrecta.")
        
    tables = [
        ("product_returns", models.DevolucionProducto),
        ("sales_history", models.HistorialVenta),
        ("cash_movements", models.MovimientoCaja),
        ("customer_payments", models.PagoCliente),
        ("purchase_items", models.ElementoCompra),
        ("purchases", models.Compra),
        ("invoices", models.Factura),
        ("shifts", models.Turno),
        ("product_variants", models.VarianteProducto),
        ("products", models.Producto),
        ("billing_profiles", models.PerfilFacturacion),
        ("customers", models.Cliente),
        ("suppliers", models.Proveedor),
        ("users", models.Usuario),
        ("store_settings", models.ConfiguracionesTienda),
        ("notifications", models.Notificacion)
    ]
    
    try:
        # Delete rows scoped to current inquilino
        for name, model in tables:
            if hasattr(model, "inquilino_id"):
                db.query(model).filter(model.inquilino_id == current_user.inquilino_id).delete()
            elif model.__tablename__ == "purchase_items":
                # Eliminar elementos de compras asociadas al inquilino
                purchases_subquery = db.query(models.Compra.id).filter(models.Compra.inquilino_id == current_user.inquilino_id).subquery()
                db.query(models.ElementoCompra).filter(models.ElementoCompra.compra_id.in_(purchases_subquery)).delete(synchronize_session=False)
        db.commit()
        
        # Insert rows, overriding inquilino_id to prevent hijacking
        for name, model in reversed(tables):
            rows = data.get(name, [])
            valid_keys = model.__mapper__.columns.keys()
            for row_dict in rows:
                filtered_dict = {k: v for k, v in row_dict.items() if k in valid_keys}
                if "inquilino_id" in valid_keys:
                    filtered_dict["inquilino_id"] = current_user.inquilino_id
                instance = model(**filtered_dict)
                db.add(instance)
            db.commit()
            
            # Sincronizar secuencias
            try:
                db.execute(text(f"SELECT setval(pg_get_serial_sequence('{model.__tablename__}', 'id'), COALESCE((SELECT MAX(id)+1 FROM {model.__tablename__}), 1), false);"))
                db.commit()
            except Exception as seq_err:
                db.rollback()
                
        return {"estado": "success", "message": "Base de datos restaurada correctamente."}
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Error durante la restauración: {str(e)}")


@router.get("/reports/dashboard-details")
def get_dashboard_details(db: Session = Depends(get_db), current_user: models.Usuario = Depends(get_current_user)):
    if current_user.rol not in ["admin", "supervisor"]:
        raise HTTPException(status_code=403, detail="No tienes permisos suficientes.")
        
    from datetime import datetime as dt, timedelta
    from sqlalchemy import func
    
    # 1. Active credit balance (total owed by customers)
    total_owed = db.query(func.coalesce(func.sum(models.Cliente.saldo_actual), 0.0)).filter(models.Cliente.inquilino_id == current_user.inquilino_id).scalar()
    
    # 2. Setup dates
    local_now = dt.now()
    
    today_start = local_now.replace(hour=0, minute=0, second=0, microsecond=0)
    yesterday_start = today_start - timedelta(days=1)
    
    # Month boundaries
    this_month_start = local_now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    last_month_end = this_month_start - timedelta(seconds=1)
    last_month_start = last_month_end.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    
    # Query non-cancelled sales
    sales_query = db.query(models.HistorialVenta).filter(
        models.HistorialVenta.cancelado == False,
        models.HistorialVenta.inquilino_id == current_user.inquilino_id
    )
    
    def calculate_sales_metrics(query_obj, start_dt, end_dt):
        sales_in_range = query_obj.filter(
            models.HistorialVenta.creado_en >= start_dt.isoformat(),
            models.HistorialVenta.creado_en <= end_dt.isoformat()
        ).all()
        
        revenue = 0.0
        cost = 0.0
        for s in sales_in_range:
            precio = s.price_sold
            cost_val = s.cost_price_sold if s.cost_price_sold is not None else 0.0
            revenue += (precio * s.cantidad) - s.discount
            cost += cost_val * s.cantidad
            
        profit = revenue - cost
        return revenue, profit
        
    # Sales/Profit for Today
    today_revenue, today_profit = calculate_sales_metrics(sales_query, today_start, local_now)
    # Sales/Profit for Yesterday
    yesterday_revenue, yesterday_profit = calculate_sales_metrics(sales_query, yesterday_start, today_start - timedelta(seconds=1))
    
    # Sales/Profit for This Month
    this_month_revenue, this_month_profit = calculate_sales_metrics(sales_query, this_month_start, local_now)
    # Sales/Profit for Last Month
    last_month_revenue, last_month_profit = calculate_sales_metrics(sales_query, last_month_start, last_month_end)
    
    # 3. 30-day time-series
    time_series = []
    for i in range(29, -1, -1):
        day_start = today_start - timedelta(days=i)
        day_end = day_start + timedelta(hours=23, minutes=59, seconds=59)
        
        day_rev, day_prof = calculate_sales_metrics(sales_query, day_start, day_end)
        time_series.append({
            "date": day_start.strftime("%Y-%m-%d"),
            "display_date": day_start.strftime("%d/%m"),
            "Ventas": round(day_rev, 2),
            "Utilidad": round(day_prof, 2)
        })
        
    # 4. Payment method breakdown (last 30 days)
    last_30_days_start = today_start - timedelta(days=30)
    sales_30_days = sales_query.filter(
        models.HistorialVenta.creado_en >= last_30_days_start.isoformat()
    ).all()
    
    pm_cash = 0.0
    pm_card = 0.0
    pm_credit = 0.0
    
    for s in sales_30_days:
        total = (s.price_sold * s.cantidad) - s.discount
        if s.payment_method == "efectivo":
            pm_cash += total
        elif s.payment_method == "tarjeta":
            pm_card += total
        elif s.payment_method == "credito":
            pm_credit += total
        elif s.payment_method == "mixto":
            pm_cash += s.cash_amount
            pm_card += s.card_amount
            
    pm_total = pm_cash + pm_card + pm_credit
    
    return {
        "today": {
            "revenue": round(today_revenue, 2),
            "profit": round(today_profit, 2),
            "yesterday_revenue": round(yesterday_revenue, 2),
            "yesterday_profit": round(yesterday_profit, 2)
        },
        "month": {
            "revenue": round(this_month_revenue, 2),
            "profit": round(this_month_profit, 2),
            "last_month_revenue": round(last_month_revenue, 2),
            "last_month_profit": round(last_month_profit, 2)
        },
        "credit": {
            "total_owed": round(total_owed, 2)
        },
        "time_series": time_series,
        "payment_methods": {
            "cash": round(pm_cash, 2),
            "card": round(pm_card, 2),
            "credit": round(pm_credit, 2),
            "total": round(pm_total, 2)
        }
    }


# ==========================================
# WEBHOOK DE WHATSAPP
# ==========================================

@router.get("/webhook/whatsapp")
def verificar_webhook(request: Request):
    """
    Meta llama a esta ruta GET para validar que tu servidor está en línea
    y que tienes el token secreto correcto.
    """
    params = request.query_params
    token_verificacion = "MiTokenSecretoDeWhatsApp123"  # Este token lo defines tú en el portal de Meta
    
    hub_mode = params.get("hub.mode")
    hub_verify_token = params.get("hub.verify_token")
    hub_challenge = params.get("hub.challenge")
    
    if hub_mode == "subscribe" and hub_verify_token == token_verificacion:
        # Debemos responder con el challenge en texto plano
        return Response(content=hub_challenge, media_type="text/plain")
        
    return Response(content="Token de verificación incorrecto", status_code=403)


@router.post("/webhook/whatsapp")
async def recibir_mensaje_whatsapp(request: Request, db: Session = Depends(get_db)):
    """
    Aquí llegarán los mensajes reales de WhatsApp enviados por los usuarios.
    Soporta tanto la API oficial de Meta (Cloud API) como el Easy API de OpenWA.
    Detecta el patrón 'agregar producto: NOMBRE, precio: X, cantidad: Y'
    y lo registra en la base de datos.
    """
    try:
        payload = await request.json()
        print("Mensaje de WhatsApp recibido:", payload)
        
        message_text = None
        sender = None
        is_openwa = False
        
        # 1. Detectar si el payload es de OpenWA Easy API
        if "event" in payload and payload["event"] == "message" and "data" in payload:
            data = payload["data"]
            message_text = data.get("body")
            sender = data.get("from")
            is_openwa = True
            print(f"Mensaje de OpenWA de {sender}: {message_text}")
            
        # 2. Detectar si es de la API oficial de Meta (Cloud API)
        elif "entry" in payload:
            for entry in payload["entry"]:
                for change in entry.get("changes", []):
                    value = change.get("value", {})
                    if "messages" in value:
                        for message in value["messages"]:
                            if "text" in message:
                                message_text = message["text"]["body"]
                                sender = message.get("from")
                                print(f"Mensaje de Meta de {sender}: {message_text}")
                                
        if message_text:
            text_lower = message_text.lower().strip()
            
            if text_lower.startswith("agregar producto:"):
                # Extraer los datos del producto
                # Formato esperado: "agregar producto: Nombre, precio: X, cantidad: Y"
                datos_str = message_text[len("agregar producto:"):].strip()
                
                # Separar por comas
                partes = [p.strip() for p in datos_str.split(",")]
                if len(partes) >= 1:
                    nombre = partes[0]
                    precio = 0.0
                    cantidad = 0
                    
                    # Buscar precio y cantidad en las partes restantes
                    for parte in partes[1:]:
                        parte_lower = parte.lower()
                        if parte_lower.startswith("precio:"):
                            try:
                                precio = float(parte_lower.replace("precio:", "").strip())
                            except ValueError:
                                pass
                        elif parte_lower.startswith("cantidad:") or parte_lower.startswith("stock:"):
                            try:
                                cantidad = int(parte_lower.replace("cantidad:", "").replace("stock:", "").strip())
                            except ValueError:
                                pass
                    
                    # Registrar en la base de datos
                    nuevo_prod = models.Producto(
                        name=nombre,
                        precio=precio,
                        cantidad=cantidad,
                        clave_sat="01010101",
                        clave_unidad_sat="H87"
                    )
                    db.add(nuevo_prod)
                    db.commit()
                    db.refresh(nuevo_prod)
                    
                    success_msg = f"Producto '{nombre}' agregado exitosamente via WhatsApp con precio={precio} y cantidad={cantidad}."
                    print(success_msg)
                    
                    # Si viene de OpenWA, enviamos confirmación automática
                    if is_openwa and sender:
                        send_whatsapp_message(
                            sender, 
                            f"✅ *Producto registrado con éxito*\n\n"
                            f"📦 *Nombre:* {nombre}\n"
                            f"💵 *Precio:* ${precio:,.2f}\n"
                            f"🔢 *Stock inicial:* {cantidad}"
                        )
                        
    except Exception as e:
        print(f"Error procesando el Webhook de WhatsApp: {e}")
        
    return {"estado": "recibido"}


# ==========================================
# ENVÍO DE TICKETS Y FACTURAS POR WHATSAPP
# ==========================================

@router.post("/sales/{venta_id}/whatsapp")
def send_sale_ticket_whatsapp(
    venta_id: int,
    req_body: schemas.WhatsAppPeticionEnvio,
    db: Session = Depends(get_db),
    current_user: models.Usuario = Depends(get_current_user)
):
    """
    Formatea el ticket de compra y lo envía al número de WhatsApp especificado usando OpenWA.
    """
    ref_sale = db.query(models.HistorialVenta).filter(models.HistorialVenta.id == venta_id).first()
    if not ref_sale:
        raise HTTPException(status_code=404, detail="Venta no encontrada")
        
    sales = db.query(
        models.HistorialVenta.id,
        models.HistorialVenta.producto_id,
        models.HistorialVenta.variante_id,
        models.HistorialVenta.cantidad,
        models.HistorialVenta.price_sold,
        models.HistorialVenta.discount,
        models.HistorialVenta.creado_en,
        models.HistorialVenta.payment_method,
        models.HistorialVenta.cash_amount,
        models.HistorialVenta.card_amount,
        models.Producto.nombre.label("nombre_producto")
    ).join(
        models.Producto, models.Producto.id == models.HistorialVenta.producto_id
    ).filter(
        models.HistorialVenta.creado_en == ref_sale.creado_en
    ).all()

    if not sales:
        raise HTTPException(status_code=404, detail="No se encontraron artículos para esta venta")

    settings = db.query(models.ConfiguracionesTienda).filter(models.ConfiguracionesTienda.id == 1).first()
    store_settings_dict = {
        "nombre_tienda": settings.nombre_tienda if settings else "ABARROTES ED & E",
        "direccion": settings.direccion if settings else "",
        "telefono": settings.telefono if settings else "",
        "pie_ticket": settings.pie_ticket if settings else "¡Gracias por su compra!"
    }

    customer_name = None
    if ref_sale.cliente_id:
        cust = db.query(models.Cliente).filter(models.Cliente.id == ref_sale.cliente_id).first()
        if cust:
            customer_name = cust.nombre

    cashier_name = "N/A"
    if ref_sale.usuario_id:
        cashier = db.query(models.Usuario).filter(models.Usuario.id == ref_sale.usuario_id).first()
        if cashier:
            cashier_name = cashier.nombre_completo

    elementos = []
    subtotal = 0.0
    discount_total = 0.0

    for s in sales:
        precio = s.price_sold or 0.0
        item_subtotal = precio * s.cantidad
        subtotal += item_subtotal
        discount_total += s.discount or 0.0
        elementos.append({
            "nombre_producto": s.nombre_producto,
            "cantidad": s.cantidad,
            "precio": precio,
            "discount": s.discount
        })

    tasa_impuesto = settings.tasa_impuesto if settings else 16.0
    tax_factor = 1 + (tasa_impuesto / 100)
    total = round(subtotal - discount_total, 2)
    subtotal_no_tax = round((subtotal - discount_total) / tax_factor, 2)

    cash_paid = sum(s.cash_amount for s in sales)
    card_paid = sum(s.card_amount for s in sales)

    sale_data = {
        "id": venta_id,
        "creado_en": ref_sale.creado_en,
        "cashier": cashier_name,
        "customer_name": customer_name,
        "elementos": elementos,
        "subtotal": subtotal_no_tax,
        "discount": discount_total,
        "total": total,
        "payment_method": ref_sale.payment_method,
        "cash_amount": cash_paid,
        "card_amount": card_paid
    }

    msg = format_ticket_message(sale_data, store_settings_dict)
    success = send_whatsapp_message(req_body.phone_number, msg)
    if not success:
        raise HTTPException(status_code=500, detail="No se pudo enviar el ticket por WhatsApp. Verifique la conexión con OpenWA.")
        
    return {"estado": "success", "message": "Ticket enviado exitosamente por WhatsApp."}


@router.post("/billing/invoices/{factura_id}/whatsapp")
def send_invoice_whatsapp(
    factura_id: int,
    req_body: schemas.WhatsAppPeticionEnvio,
    request: Request,
    db: Session = Depends(get_db),
    current_user: models.Usuario = Depends(get_current_user)
):
    """
    Envía los detalles de la factura y los enlaces de descarga PDF/XML vía WhatsApp usando OpenWA.
    """
    inv = db.query(models.Factura).filter(models.Factura.id == factura_id).first()
    if not inv:
        raise HTTPException(status_code=404, detail="Factura no encontrada")

    settings = db.query(models.ConfiguracionesTienda).filter(models.ConfiguracionesTienda.id == 1).first()
    nombre_tienda = settings.nombre_tienda if settings else "ABARROTES ED & E"

    base_url = str(request.base_url).rstrip("/")
    # Construir enlaces de descarga basados en el host actual
    xml_link = f"{base_url}/api/billing/invoices/{inv.uuid}/xml"
    pdf_link = f"{base_url}/api/billing/invoices/{inv.uuid}/pdf"

    msg_lines = [
        "🧾 *FACTURA ELECTRÓNICA CFDI 4.0*",
        f"🏢 *{nombre_tienda}*",
        "-----------------------------------------",
        f"🆔 *Folio Fiscal (UUID):*",
        f"_{inv.uuid}_",
        f"📅 *Fecha Certificación:* {inv.creado_en}",
        f"💵 *Monto Total:* ${inv.monto_total:,.2f} MXN",
        f"📊 *Estado:* {'Vigente' if inv.estado == 'active' else 'Cancelado'}",
        "-----------------------------------------",
        "📄 *Descargar PDF:*",
        pdf_link,
        "",
        "🔗 *Descargar XML:*",
        xml_link,
        "-----------------------------------------",
        "¡Gracias por su preferencia!"
    ]

    msg = "\n".join(msg_lines)
    success = send_whatsapp_message(req_body.phone_number, msg)
    if not success:
        raise HTTPException(status_code=500, detail="No se pudo enviar la factura por WhatsApp.")
        
    return {"estado": "success", "message": "Factura enviada exitosamente por WhatsApp."}

# --- SUPERADMIN SAAS CONTROL PANEL ENDPOINTS ---

def check_superadmin_privilege(current_user: models.Usuario):
    # Verify current_user belongs to inquilino_id 1 (creator store) and has admin rol
    if current_user.inquilino_id != 1 or current_user.rol != 'admin':
        raise HTTPException(
            status_code=estado.HTTP_403_FORBIDDEN,
            detail="Acceso denegado. Se requieren privilegios de Superadministrador del SaaS."
        )

@router.get("/superadmin/tenants", response_model=List[schemas.SuperAdminInquilinoResponse])
def superadmin_get_tenants(db: Session = Depends(get_db), current_user: models.Usuario = Depends(get_current_user)):
    check_superadmin_privilege(current_user)
    
    tenants = db.query(models.Inquilino).all()
    results = []
    
    for t in tenants:
        # Get admin user info for this inquilino
        admin_user = db.query(models.Usuario).filter(
            models.Usuario.inquilino_id == t.id,
            models.Usuario.rol == 'admin'
        ).first()
        
        # Calculate counts
        product_count = db.query(models.Producto).filter(models.Producto.inquilino_id == t.id).count()
        sale_count = db.query(models.HistorialVenta).filter(models.HistorialVenta.inquilino_id == t.id).count()
        
        results.append({
            "id": t.id,
            "nombre": t.nombre,
            "subdominio": t.subdominio,
            "estado_suscripcion": t.estado_suscripcion,
            "nivel_plan": t.nivel_plan,
            "creado_en": t.creado_en or datetime.utcnow().isoformat(),
            "usuario_admin": admin_user.nombre_usuario if admin_user else "N/A",
            "nombre_admin": admin_user.nombre_completo if admin_user else "N/A",
            "cantidad_productos": product_count,
            "cantidad_ventas": sale_count,
            "fecha_ultimo_pago": t.fecha_ultimo_pago,
            "fin_suscripcion": t.fin_suscripcion
        })
        
    return results

@router.put("/superadmin/tenants/{inquilino_id}/plan", response_model=schemas.InquilinoResponse)
def superadmin_update_tenant_plan(inquilino_id: int, req: schemas.SuperAdminInquilinoUpdate, db: Session = Depends(get_db), current_user: models.Usuario = Depends(get_current_user)):
    check_superadmin_privilege(current_user)
    
    if inquilino_id == 1:
        raise HTTPException(status_code=400, detail="No se puede modificar la suscripción del inquilino principal.")
        
    inquilino = db.query(models.Inquilino).filter(models.Inquilino.id == inquilino_id).first()
    if not inquilino:
        raise HTTPException(status_code=404, detail="Tienda no encontrada.")
        
    inquilino.nivel_plan = req.nivel_plan
    inquilino.estado_suscripcion = req.estado_suscripcion
    db.commit()
    db.refresh(inquilino)
    return inquilino

@router.put("/superadmin/tenants/{inquilino_id}/reset-password")
def superadmin_reset_tenant_password(inquilino_id: int, req: schemas.SuperAdminInquilinoResetPassword, db: Session = Depends(get_db), current_user: models.Usuario = Depends(get_current_user)):
    check_superadmin_privilege(current_user)
    
    # Buscar el administrador de ese inquilino
    admin_user = db.query(models.Usuario).filter(
        models.Usuario.inquilino_id == inquilino_id,
        models.Usuario.rol == 'admin'
    ).first()
    if not admin_user:
        raise HTTPException(status_code=404, detail="Usuario administrador no encontrado para esta tienda.")
    
    # Actualizar contraseña
    admin_user.contrasena_encriptada = auth.get_password_hash(req.new_password)
    db.commit()

    # Log action
    log_entry = models.BitacoraUsuario(
        inquilino_id=inquilino_id,
        usuario=admin_user.nombre_usuario,
        nombre_completo=admin_user.nombre_completo,
        rol=admin_user.rol,
        accion="actualizacion",
        detalles="Contraseña del administrador restablecida por el Superadministrador desde la Consola SaaS.",
        fecha_hora=datetime.utcnow().isoformat()
    )
    db.add(log_entry)
    db.commit()
    
    return {"estado": "success", "message": f"Contraseña restablecida exitosamente para el usuario admin @{admin_user.nombre_usuario}."}

@router.delete("/superadmin/tenants/{inquilino_id}")
def superadmin_delete_tenant(inquilino_id: int, db: Session = Depends(get_db), current_user: models.Usuario = Depends(get_current_user)):
    check_superadmin_privilege(current_user)
    
    if inquilino_id == 1:
        raise HTTPException(status_code=400, detail="No se puede eliminar el inquilino principal.")
        
    inquilino = db.query(models.Inquilino).filter(models.Inquilino.id == inquilino_id).first()
    if not inquilino:
        raise HTTPException(status_code=404, detail="Tienda no encontrada.")
        
    # We clean up child tables first
    tables_to_clean = [
        ("elementos_compra", "compra_id IN (SELECT id FROM compras WHERE inquilino_id = :tid)"),
        ("compras", "inquilino_id = :tid"),
        ("movimientos_caja", "inquilino_id = :tid"),
        ("pagos_cliente", "inquilino_id = :tid"),
        ("historial_ventas", "inquilino_id = :tid"),
        ("devoluciones_producto", "inquilino_id = :tid"),
        ("variantes_producto", "inquilino_id = :tid"),
        ("productos", "inquilino_id = :tid"),
        ("perfiles_facturacion", "inquilino_id = :tid"),
        ("facturas", "inquilino_id = :tid"),
        ("proveedores", "inquilino_id = :tid"),
        ("clientes", "inquilino_id = :tid"),
        ("notificaciones", "inquilino_id = :tid"),
        ("turnos", "inquilino_id = :tid"),
        ("configuraciones_tienda", "inquilino_id = :tid"),
        ("usuarios", "inquilino_id = :tid"),
        ("bitacora_usuarios", "inquilino_id = :tid"),
    ]
    
    try:
        from sqlalchemy import text
        for table, condition in tables_to_clean:
            db.execute(text(f"DELETE FROM {table} WHERE {condition}"), {"tid": inquilino_id})
            
        db.delete(inquilino)
        db.commit()
        return {"estado": "success", "message": f"Inquilino {inquilino_id} y todos sus datos relacionados fueron eliminados."}
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Error al eliminar inquilino: {str(e)}")

# --- STRIPE BILLING & AUTO SUBSCRIPTION AUTOMATION ---

import stripe

@router.post("/billing/create-checkout-session")
def create_checkout_session(db: Session = Depends(get_db), current_user: models.Usuario = Depends(get_current_user)):
    if not current_user.inquilino_id:
        raise HTTPException(status_code=400, detail="El usuario no pertenece a ninguna tienda.")
        
    inquilino = db.query(models.Inquilino).filter(models.Inquilino.id == current_user.inquilino_id).first()
    if not inquilino:
        raise HTTPException(status_code=404, detail="Tienda no encontrada.")

    stripe_key = os.getenv("STRIPE_SECRET_KEY")
    if not stripe_key:
        # Fallback to simulation mode for testing out of the box!
        simulation_url = f"/payment-simulation?inquilino_id={inquilino.id}&nombre_tienda={inquilino.nombre}"
        return {"url": simulation_url, "simulated": True}

    # Real Stripe session setup
    stripe.api_key = stripe_key
    try:
        # Create or retrieve customer
        cliente_id = inquilino.stripe_id_cliente
        if not cliente_id:
            customer = stripe.Customer.create(
                name=inquilino.nombre,
                metadata={"inquilino_id": inquilino.id}
            )
            cliente_id = customer.id
            inquilino.stripe_id_cliente = cliente_id
            db.commit()

        # Build callback URLs
        origin = os.getenv("FRONTEND_URL", "http://localhost:5173")
        
        # Stripe Checkout Session
        session = stripe.checkout.Session.create(
            customer=cliente_id,
            payment_method_types=['card'],
            line_items=[{
                'price_data': {
                    'currency': 'mxn',
                    'product_data': {
                        'nombre': f"Suscripción Mensual - {inquilino.nombre}",
                        'description': 'Acceso completo e ilimitado al sistema de Punto de Venta.',
                    },
                    'unit_amount': 49900, # $499.00 MXN in cents
                    'recurring': {
                        'interval': 'month',
                    },
                },
                'cantidad': 1,
            }],
            mode='subscription',
            success_url=f"{origin}/dashboard?payment=success",
            cancel_url=f"{origin}/dashboard?payment=cancel",
            metadata={"inquilino_id": inquilino.id}
        )
        return {"url": session.url, "simulated": False}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error al crear sesión de Stripe: {str(e)}")


@router.post("/billing/webhook")
async def stripe_webhook(request: Request, db: Session = Depends(get_db)):
    payload = await request.body()
    sig_header = request.headers.get("stripe-signature")
    endpoint_secret = os.getenv("STRIPE_WEBHOOK_SECRET")
    stripe_key = os.getenv("STRIPE_SECRET_KEY")

    if not sig_header or not endpoint_secret or not stripe_key:
        raise HTTPException(status_code=400, detail="Configuración de webhook inválida o incompleta.")

    stripe.api_key = stripe_key
    try:
        event = stripe.Webhook.construct_event(
            payload, sig_header, endpoint_secret
        )
    except ValueError:
        raise HTTPException(status_code=400, detail="Payload inválido")
    except stripe.error.SignatureVerificationError:
        raise HTTPException(status_code=400, detail="Firma de webhook inválida")

    event_type = event['tipo']
    
    if event_type in ["checkout.session.completed", "invoice.payment_succeeded"]:
        session = event['data']['object']
        
        # Extract inquilino_id
        inquilino_id = None
        if 'metadata' in session and 'inquilino_id' in session['metadata']:
            inquilino_id = int(session['metadata']['inquilino_id'])
        elif 'customer' in session:
            # Look up customer in tenants
            cust = db.query(models.Inquilino).filter(models.Inquilino.stripe_id_cliente == session['customer']).first()
            if cust:
                inquilino_id = cust.id

        if inquilino_id:
            inquilino = db.query(models.Inquilino).filter(models.Inquilino.id == inquilino_id).first()
            if inquilino:
                inquilino.estado_suscripcion = "active"
                inquilino.nivel_plan = "premium"
                inquilino.fecha_ultimo_pago = datetime.utcnow().isoformat()
                # Set subscription end to 30 days from now
                inquilino.fin_suscripcion = (datetime.utcnow() + timedelta(days=30)).isoformat()
                
                # Update stripe subscription id if present
                if 'subscription' in session:
                    inquilino.stripe_id_suscripcion = session['subscription']
                    
                db.commit()
                print(f"Suscripción automatizada con éxito para Tenant {inquilino_id} a través de Stripe.")
                
    elif event_type in ["invoice.payment_failed", "customer.subscription.deleted"]:
        obj = event['data']['object']
        cliente_id = obj.get("customer")
        if cliente_id:
            inquilino = db.query(models.Inquilino).filter(models.Inquilino.stripe_id_cliente == cliente_id).first()
            if inquilino:
                inquilino.estado_suscripcion = "suspended"
                db.commit()
                print(f"Suscripción del Tenant {inquilino.id} suspendida por fallo en pago de Stripe.")

    return {"estado": "success"}


@router.post("/billing/simulate-success")
def simulate_payment_success(req: schemas.SimulatePaymentRequest = None, db: Session = Depends(get_db), current_user: models.Usuario = Depends(get_current_user)):
    # Anyone can simulate for their own inquilino
    if not current_user.inquilino_id:
        raise HTTPException(status_code=400, detail="El usuario no tiene tienda vinculada.")

    inquilino = db.query(models.Inquilino).filter(models.Inquilino.id == current_user.inquilino_id).first()
    if not inquilino:
        raise HTTPException(status_code=404, detail="Tienda no encontrada.")

    inquilino.estado_suscripcion = "active"
    inquilino.nivel_plan = "premium"
    inquilino.fecha_ultimo_pago = datetime.utcnow().isoformat()
    # Expire in 30 days
    inquilino.fin_suscripcion = (datetime.utcnow() + timedelta(days=30)).isoformat()
    
    if req:
        inquilino.metodo_pago_guardado = req.metodo_pago
        inquilino.tarjeta_marca = req.tarjeta_marca
        inquilino.tarjeta_ultimos4 = req.tarjeta_ultimos4
        inquilino.tarjeta_titular = req.tarjeta_titular
        inquilino.tarjeta_vencimiento = req.tarjeta_vencimiento

    # Log action in BitacoraUsuario
    log_entry = models.BitacoraUsuario(
        inquilino_id=inquilino.id,
        usuario=current_user.nombre_usuario,
        nombre_completo=current_user.nombre_completo,
        rol=current_user.rol,
        accion="actualizacion",
        detalles=f"Suscripción Premium activada/renovada mediante pago simulado ({req.metodo_pago if req else 'desconocido'}). Vencimiento establecido para: {inquilino.fin_suscripcion}.",
        fecha_hora=datetime.utcnow().isoformat()
    )
    db.add(log_entry)
    db.commit()
    db.refresh(inquilino)
    return {
        "estado": "success",
        "message": "Pago simulado con éxito. Tu cuenta ahora es Premium y está activa.",
        "fin_suscripcion": inquilino.fin_suscripcion
    }








